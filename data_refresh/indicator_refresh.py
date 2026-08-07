"""Indicator Refresh MVP: CSV → staging → validate → merge por columnas → ETL 007."""

from __future__ import annotations

import csv
import io
import logging
import re
from typing import Any, Dict, List, Optional, Sequence, Tuple

from column_resolver import clear_column_cache, resolve_column
from data_refresh.jobs_store import (
    get_job,
    insert_job,
    list_recent_jobs,
    new_job_id,
    update_job,
)
from data_refresh.names import STAGING_SCHEMA, assert_job_id
from database import get_db
from indicators_catalog_loader import load_indicators_catalog_raw
from tables import SCHEMA, T_TAB_MUNICIPAL, T_TAB_NACIONAL

logger = logging.getLogger(__name__)

CVE_RE = re.compile(r"^[0-9]{3}$")
ENT_RE = re.compile(r"^[0-9]{2}$")
IDENT_RE = re.compile(r"^[a-z][a-z0-9_]{0,62}$")
SPECIAL_NOMS = frozenset({"nacional", "estatal"})


def _normalize_cve_mun(raw: Any) -> str:
    """Normaliza claves municipales a 3 dígitos (001…).

    Excel suele quitar ceros a la izquierda (1, 42) o devolver float (1.0).
    El merge exige exactamente ^[0-9]{3}$ como en producción.
    """
    if raw is None:
        return ""
    if isinstance(raw, float):
        if raw != raw:  # NaN
            return ""
        if raw == int(raw):
            raw = int(raw)
        else:
            return ""
    if isinstance(raw, int):
        s = str(raw)
    else:
        s = str(raw).strip()
        if not s or s.lower() in ("none", "null", "nan"):
            return ""
        # Excel texto raro: ="001" o '001
        if s.startswith("="):
            s = s.lstrip("=").strip().strip('"').strip("'")
        s = s.strip("'").strip('"')
        if re.fullmatch(r"[0-9]+\.0+", s):
            s = s.split(".", 1)[0]
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    if len(digits) > 3:
        # Conservar últimos 3 si viniera cvegeo-like; no truncar mal
        if len(digits) == 5 and digits.startswith("12"):
            digits = digits[-3:]
        elif len(digits) > 3:
            return ""
    return digits.zfill(3)


def _normalize_ent(raw: Any) -> str:
    """Normaliza clave de entidad a 2 dígitos (01…32). Nunca zfill(3)."""
    if raw is None:
        return ""
    if isinstance(raw, float):
        if raw != raw:
            return ""
        if raw == int(raw):
            raw = int(raw)
        else:
            return ""
    if isinstance(raw, int):
        s = str(raw)
    else:
        s = str(raw).strip()
        if not s or s.lower() in ("none", "null", "nan"):
            return ""
        if s.startswith("="):
            s = s.lstrip("=").strip().strip('"').strip("'")
        s = s.strip("'").strip('"')
        if re.fullmatch(r"[0-9]+\.0+", s):
            s = s.split(".", 1)[0]
    digits = re.sub(r"\D", "", s)
    if not digits:
        return ""
    try:
        n = int(digits)
    except ValueError:
        return ""
    if n < 1 or n > 32:
        return ""
    return str(n).zfill(2)


def _template_is_nacional(template: Dict[str, Any]) -> bool:
    tgt = (
        template.get("target_table")
        or template.get("table")
        or template.get("scope")
        or ""
    )
    return str(tgt).lower() in (T_TAB_NACIONAL, "nacional", f"atlas.{T_TAB_NACIONAL}")


def indicator_staging_name(job_id: str) -> str:
    return f"ir_{assert_job_id(job_id)}"[:63]


def _qident(name: str) -> str:
    n = (name or "").strip()
    if not IDENT_RE.match(n.lower()):
        raise ValueError(f"Identificador inválido: {name!r}")
    return n


def _drop_staging(table: str) -> None:
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(f'DROP TABLE IF EXISTS "{STAGING_SCHEMA}"."{table}" CASCADE')
        conn.commit()


def ensure_staging_schema() -> None:
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(f"CREATE SCHEMA IF NOT EXISTS {STAGING_SCHEMA}")
        conn.commit()


def list_indicator_templates() -> List[Dict[str, Any]]:
    """Plantillas: una entrada por (indicador × tabla destino municipal|nacional)."""
    from data_refresh.indicator_nacional import build_nacional_template_entry

    cat = load_indicators_catalog_raw()
    out: List[Dict[str, Any]] = []
    with get_db() as conn:
        for ind in cat.get("indicators") or []:
            if not ind.get("enabled", True):
                continue
            src = ind.get("source") or {}
            primary = str(src.get("primary_table") or "").lower()
            fields = ind.get("fields") or []

            mun_metrics: List[Dict[str, Any]] = []
            for f in fields:
                st = str(f.get("source_table") or "tab_municipal").lower()
                if st not in ("tab_municipal", ""):
                    continue
                col = str(f.get("column") or f.get("key") or "").strip()
                if not col:
                    continue
                aliases = [col] + list(f.get("column_aliases") or [])
                resolved = resolve_column(conn, SCHEMA, T_TAB_MUNICIPAL, aliases)
                mun_metrics.append(
                    {
                        "key": f.get("key") or col,
                        "column": col,
                        "resolved_column": resolved,
                        "aliases": aliases,
                        "label": f.get("label") or col,
                        "type": f.get("type") or "float",
                        "in_database": bool(resolved),
                    }
                )
            if mun_metrics or "tab_municipal" in primary:
                if mun_metrics:
                    headers = ["cve_mun", "nom_mun"] + [
                        m["column"] for m in mun_metrics if m.get("in_database")
                    ]
                    out.append(
                        {
                            "id": ind.get("id"),
                            "indicator_id": ind.get("id"),
                            "label": ind.get("label") or ind.get("id"),
                            "group_id": ind.get("group_id"),
                            "subtitle": ind.get("subtitle") or "",
                            "table": T_TAB_MUNICIPAL,
                            "target_table": T_TAB_MUNICIPAL,
                            "scope": "municipal",
                            "key_columns": ["cve_mun", "nom_mun"],
                            "metrics": mun_metrics,
                            "csv_headers_hint": ";".join(headers),
                            "metrics_in_db": sum(
                                1 for m in mun_metrics if m.get("in_database")
                            ),
                        }
                    )

            nat_entry = build_nacional_template_entry(conn, ind, fields)
            if nat_entry:
                out.append(nat_entry)

    out.sort(
        key=lambda t: (
            (t.get("group_id") or ""),
            (t.get("label") or ""),
            0 if t.get("scope") == "municipal" else 1,
        )
    )
    return out


def _get_template(template_id: str) -> Dict[str, Any]:
    tid = (template_id or "").strip()
    for t in list_indicator_templates():
        if t.get("id") == tid:
            return t
    raise ValueError(f"TEMPLATE_NOT_FOUND:{template_id}")


def _decode_csv(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    if not content:
        raise ValueError("EMPTY_FILE")
    text = content.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=";,\t")
        delimiter = dialect.delimiter
    except csv.Error:
        delimiter = ";" if sample.count(";") >= sample.count(",") else ","
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    if not reader.fieldnames:
        raise ValueError("CSV_SIN_ENCABEZADOS")
    headers = [str(h or "").strip() for h in reader.fieldnames]
    rows: List[Dict[str, str]] = []
    for raw in reader:
        row = {
            str(k or "").strip(): ("" if v is None else str(v).strip())
            for k, v in raw.items()
            if k is not None
        }
        if not any(row.values()):
            continue
        rows.append(row)
    if not rows:
        raise ValueError("CSV_SIN_FILAS")
    return headers, rows


def _cell_to_str(val: Any) -> str:
    if val is None:
        return ""
    if isinstance(val, float):
        if val != val:
            return ""
        if val == int(val):
            return str(int(val))
        return str(val)
    if isinstance(val, int):
        return str(val)
    return str(val).strip()


def _decode_xlsx(content: bytes) -> Tuple[List[str], List[Dict[str, str]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ValueError("XLSX_NO_DISPONIBLE: openpyxl no está instalado") from exc
    if not content:
        raise ValueError("EMPTY_FILE")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration as exc:
        raise ValueError("CSV_SIN_ENCABEZADOS") from exc
    headers = [_cell_to_str(h) for h in header_row]
    if not any(headers):
        raise ValueError("CSV_SIN_ENCABEZADOS")
    # Rellenar encabezados vacíos para no perder columnas
    headers = [h if h else f"col_{i+1}" for i, h in enumerate(headers)]
    out: List[Dict[str, str]] = []
    for vals in rows_iter:
        if vals is None:
            continue
        row = {
            headers[i]: _cell_to_str(vals[i] if i < len(vals) else "")
            for i in range(len(headers))
        }
        if not any(row.values()):
            continue
        out.append(row)
    wb.close()
    if not out:
        raise ValueError("CSV_SIN_FILAS")
    return headers, out


def _decode_tabular(
    content: bytes, filename: Optional[str] = None
) -> Tuple[List[str], List[Dict[str, str]]]:
    name = (filename or "").lower()
    is_xlsx = name.endswith(".xlsx") or (
        content[:2] == b"PK" and not name.endswith(".csv")
    )
    if is_xlsx:
        return _decode_xlsx(content)
    return _decode_csv(content)


def _norm_header_map(headers: Sequence[str]) -> Dict[str, str]:
    """lower header -> original header."""
    out: Dict[str, str] = {}
    for h in headers:
        key = h.strip().lower()
        if key and key not in out:
            out[key] = h.strip()
    return out


def _find_csv_col(header_map: Dict[str, str], candidates: Sequence[str]) -> Optional[str]:
    for c in candidates:
        hit = header_map.get(str(c).strip().lower())
        if hit:
            return hit
    return None


def _parse_number(raw: str, typ: str) -> Optional[float]:
    s = (raw or "").strip()
    if s == "" or s.lower() in ("null", "none", "na", "n/d", "-"):
        return None
    s = s.replace(",", "")
    try:
        val = float(s)
    except ValueError as exc:
        raise ValueError(f"VALOR_NO_NUMERICO:{raw!r}") from exc
    if typ == "integer":
        return float(int(round(val)))
    return val


def _prod_municipal_cves() -> List[str]:
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM cve_mun::text) AS cve
                  FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                 WHERE TRIM(BOTH FROM cve_mun::text) ~ '^[0-9]{{3}}$'
                   AND LOWER(TRIM(BOTH FROM COALESCE(nom_mun::text, '')))
                       NOT IN ('nacional', 'estatal')
                 ORDER BY 1
                """
            )
            return [str(r["cve"]) for r in (cur.fetchall() or []) if r.get("cve")]


def _prod_special_noms() -> Dict[str, bool]:
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT LOWER(TRIM(BOTH FROM nom_mun::text)) AS nom
                  FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                 WHERE LOWER(TRIM(BOTH FROM nom_mun::text)) IN ('nacional', 'estatal')
                """
            )
            found = {str(r["nom"]) for r in (cur.fetchall() or []) if r.get("nom")}
    return {"nacional": "nacional" in found, "estatal": "estatal" in found}


def _build_report(
    *,
    template: Dict[str, Any],
    headers: List[str],
    rows: List[Dict[str, str]],
    mapped_metrics: List[Dict[str, Any]],
    cve_col: str,
    nom_col: Optional[str],
) -> Dict[str, Any]:
    prod_cves = set(_prod_municipal_cves())
    csv_cves: set = set()
    csv_special = {"nacional": False, "estatal": False}
    bad_cve = 0
    null_counts = {m["resolved_column"]: 0 for m in mapped_metrics}
    type_errors: List[str] = []
    collision_warnings: List[str] = []

    for row in rows:
        nom = (row.get(nom_col) or "").strip().lower() if nom_col else ""
        cve = _normalize_cve_mun(row.get(cve_col))
        is_special = nom in SPECIAL_NOMS
        if is_special:
            csv_special[nom] = True
            # Estatal suele traer cve_mun=012 (cve_ent espuria) = Ayutla; no mezclar con padrón
            if CVE_RE.match(cve) and cve in prod_cves:
                collision_warnings.append(
                    f"Fila «{nom.capitalize()}» trae cve_mun={cve}, igual que un municipio real; "
                    "el merge usará solo nom_mun (no cve)."
                )
        elif CVE_RE.match(cve):
            if cve in csv_cves:
                collision_warnings.append(f"cve_mun duplicada en archivo: {cve}")
            csv_cves.add(cve)
        else:
            bad_cve += 1

        for m in mapped_metrics:
            csv_h = m["csv_header"]
            raw = row.get(csv_h, "")
            if raw == "":
                null_counts[m["resolved_column"]] += 1
                continue
            try:
                _parse_number(raw, m.get("type") or "float")
            except ValueError:
                type_errors.append(
                    f"Fila cve={cve or nom!r} col={m['column']}: valor {raw!r}"
                )

    missing = sorted(prod_cves - csv_cves)
    extra = sorted(csv_cves - prod_cves)
    prod_special = _prod_special_noms()

    warnings: List[str] = list(dict.fromkeys(collision_warnings))  # unique, stable
    infos: List[str] = []
    if missing:
        warnings.append(
            f"Faltan {len(missing)} municipios del padrón actual "
            f"(ej. {', '.join(missing[:8])}{'…' if len(missing) > 8 else ''})."
        )
    if extra:
        warnings.append(
            f"Hay {len(extra)} cve_mun no presentes en producción "
            f"(no se insertarán filas nuevas en MVP merge)."
        )
    if bad_cve:
        warnings.append(f"{bad_cve} filas con cve_mun inválido (se ignoran en merge municipal).")
    if type_errors:
        warnings.append(
            f"{len(type_errors)} valores no numéricos "
            f"(ej. {type_errors[0]})."
        )
    for key, present_prod in prod_special.items():
        if present_prod and not csv_special.get(key):
            infos.append(
                f"Fila «{key.capitalize()}» no viene en el CSV; se conserva la de producción."
            )
        elif csv_special.get(key):
            infos.append(f"Fila «{key.capitalize()}» incluida; se actualizará por nom_mun.")

    level = "ok"
    label = "Listo para aplicar"
    dup_msgs = [w for w in warnings if "duplicada" in w.lower()]
    if type_errors or (missing and len(missing) > 10) or dup_msgs or (
        null_counts
        and any(v > max(1, len(rows) // 2) for v in null_counts.values())
    ):
        level = "warn"
        label = "Revisar advertencias"
    if not mapped_metrics:
        level = "block"
        label = "Sin columnas métricas válidas"
    if not csv_cves and not any(csv_special.values()):
        level = "block"
        label = "Sin filas municipales válidas"
    if dup_msgs and len(dup_msgs) > 5:
        level = "block"
        label = "Demasiadas claves duplicadas"

    # Deltas aproximados (muestra) contra producción para 1ª métrica
    deltas: List[Dict[str, Any]] = []
    if mapped_metrics and csv_cves:
        m0 = mapped_metrics[0]
        col = m0["resolved_column"]
        csv_h = m0["csv_header"]
        changed = 0
        sample: List[Dict[str, Any]] = []
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT TRIM(BOTH FROM cve_mun::text) AS cve,
                           "{col}" AS val
                      FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                     WHERE TRIM(BOTH FROM cve_mun::text) ~ '^[0-9]{{3}}$'
                       AND LOWER(TRIM(BOTH FROM COALESCE(nom_mun::text, '')))
                           NOT IN ('nacional', 'estatal')
                    """
                )
                prod_map = {
                    str(r["cve"]): r.get("val") for r in (cur.fetchall() or []) if r.get("cve")
                }
        for row in rows:
            nom = (row.get(nom_col) or "").strip().lower() if nom_col else ""
            if nom in SPECIAL_NOMS:
                continue  # Nacional/Estatal: no delta por cve (colisión 012)
            cve = _normalize_cve_mun(row.get(cve_col))
            if cve not in prod_cves:
                continue
            raw = row.get(csv_h, "")
            try:
                new_v = _parse_number(raw, m0.get("type") or "float")
            except ValueError:
                continue
            old_v = prod_map.get(cve)
            if old_v is None and new_v is None:
                continue
            try:
                old_f = float(old_v) if old_v is not None else None
            except (TypeError, ValueError):
                old_f = None
            if old_f != new_v:
                changed += 1
                if len(sample) < 5:
                    sample.append(
                        {
                            "cve_mun": cve,
                            "before": old_f,
                            "after": new_v,
                            "column": m0["column"],
                        }
                    )
        deltas.append(
            {
                "column": m0["column"],
                "label": m0.get("label"),
                "rows_changed": changed,
                "sample": sample,
            }
        )

    mun_ok = len(csv_cves & prod_cves)
    return {
        "kind": "indicator",
        "strategy": "column_merge_by_cve_mun",
        "strategy_label": "Actualización por columnas (merge)",
        "template_id": template.get("id"),
        "template_label": template.get("label"),
        "target_table": T_TAB_MUNICIPAL,
        "csv_headers": headers,
        "mapped_metrics": mapped_metrics,
        "row_count_csv": len(rows),
        "municipal_in_csv": len(csv_cves),
        "municipal_matched": mun_ok,
        "municipal_missing": missing,
        "municipal_extra": extra,
        "special_rows_csv": csv_special,
        "null_counts": null_counts,
        "type_error_count": len(type_errors),
        "type_error_samples": type_errors[:8],
        "deltas": deltas,
        "warnings": warnings,
        "infos": infos
        + [
            "MVP: solo se actualizan columnas del CSV; el resto de tab_municipal se conserva.",
            "No se insertan municipios nuevos (solo merge por cve_mun / Nacional / Estatal).",
        ],
        "validation": {
            "level": level,
            "label": label,
            "checks": {
                "template": True,
                "metrics": bool(mapped_metrics),
                "municipalities": mun_ok > 0,
                "types": len(type_errors) == 0,
                "duplicates": len(dup_msgs) == 0,
                "swap_available": level != "block",
            },
            "checklist": [
                {
                    "id": "metrics",
                    "label": "Columnas / métricas mapeadas",
                    "status": "ok" if mapped_metrics else "block",
                    "detail": f"{len(mapped_metrics)} cols",
                },
                {
                    "id": "municipalities",
                    "label": "Claves cve_mun válidas",
                    "status": "ok" if mun_ok > 0 else "block",
                    "detail": f"{mun_ok} matched",
                },
                {
                    "id": "duplicates",
                    "label": "Duplicados de clave",
                    "status": "warn" if dup_msgs else "ok",
                    "detail": f"{len(dup_msgs)} avisos" if dup_msgs else "OK",
                },
                {
                    "id": "nulls",
                    "label": "Nulos en columnas críticas",
                    "status": (
                        "warn"
                        if any(v > max(1, len(rows) // 2) for v in null_counts.values())
                        else "ok"
                    ),
                    "detail": "revisar" if null_counts else "OK",
                },
                {
                    "id": "types",
                    "label": "Tipos numéricos",
                    "status": "warn" if type_errors else "ok",
                    "detail": f"{len(type_errors)} errores",
                },
            ],
        },
        "summary_line": (
            f"{mun_ok} municipios OK"
            + (f", {len(missing)} faltantes" if missing else ", 0 faltantes")
        ),
        "phase": "ready",
        "progress": 100,
        "label": "Validación lista",
    }


def create_indicator_job_from_csv(
    *,
    content: bytes,
    filename: str,
    template_id: str,
    user_id: Optional[int] = None,
) -> Dict[str, Any]:
    template = _get_template(template_id)
    if _template_is_nacional(template):
        from data_refresh.indicator_nacional import create_nacional_job

        return create_nacional_job(
            content=content,
            filename=filename,
            template=template,
            user_id=user_id,
        )

    headers, rows = _decode_tabular(content, filename)
    header_map = _norm_header_map(headers)

    cve_col = _find_csv_col(header_map, ("cve_mun", "cve", "municipio_cve", "CVE_MUN"))
    if not cve_col:
        raise ValueError("CSV_FALTA_CVE_MUN: se requiere columna cve_mun")
    nom_col = _find_csv_col(header_map, ("nom_mun", "municipio", "nombre", "NOM_MUN"))

    mapped_metrics: List[Dict[str, Any]] = []
    for m in template.get("metrics") or []:
        if not m.get("in_database") or not m.get("resolved_column"):
            continue
        csv_h = _find_csv_col(
            header_map,
            [m["column"], m.get("key") or "", *(m.get("aliases") or [])],
        )
        if not csv_h:
            continue
        mapped_metrics.append({**m, "csv_header": csv_h})

    if not mapped_metrics:
        raise ValueError(
            "CSV_SIN_METRICAS: ninguna columna de la plantilla está en el archivo "
            f"(esperadas: {template.get('csv_headers_hint')})"
        )

    report = _build_report(
        template=template,
        headers=headers,
        rows=rows,
        mapped_metrics=mapped_metrics,
        cve_col=cve_col,
        nom_col=nom_col,
    )
    report["filename"] = filename or "upload.csv"
    report["scope"] = "municipal"
    report["target_table"] = T_TAB_MUNICIPAL

    if report["validation"]["level"] == "block":
        # Aún creamos job failed para auditoría
        job_id = new_job_id()
        stg = indicator_staging_name(job_id)
        insert_job(
            job_id=job_id,
            user_id=user_id,
            target_table=T_TAB_MUNICIPAL,
            staging_table=stg,
            status="failed",
            report=report,
            error_message=report["validation"]["label"],
        )
        raise ValueError(report["validation"]["label"])

    ensure_staging_schema()
    job_id = new_job_id()
    stg = indicator_staging_name(job_id)
    _drop_staging(stg)

    # Crear staging estrecho + cargar
    col_defs = ['"cve_mun" text', '"nom_mun" text']
    metric_cols = []
    for m in mapped_metrics:
        rc = _qident(m["resolved_column"])
        metric_cols.append(rc)
        col_defs.append(f'"{rc}" double precision')

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f'CREATE TABLE "{STAGING_SCHEMA}"."{stg}" ({", ".join(col_defs)})'
            )
            for row in rows:
                nom = (row.get(nom_col) or "").strip() if nom_col else ""
                nom_l = nom.lower()
                is_special = nom_l in SPECIAL_NOMS
                cve = _normalize_cve_mun(row.get(cve_col))
                if is_special:
                    # Forzar merge solo por nombre: no contaminar join por cve_mun
                    # (Estatal.012 colisiona con Ayutla de los Libres).
                    cve = ""
                elif not CVE_RE.match(cve):
                    continue
                values: List[Any] = [cve or None, nom or None]
                for m in mapped_metrics:
                    raw = row.get(m["csv_header"], "")
                    try:
                        values.append(_parse_number(raw, m.get("type") or "float"))
                    except ValueError:
                        values.append(None)
                placeholders = ", ".join(["%s"] * len(values))
                cols_sql = ", ".join(
                    ['"cve_mun"', '"nom_mun"'] + [f'"{c}"' for c in metric_cols]
                )
                cur.execute(
                    f'INSERT INTO "{STAGING_SCHEMA}"."{stg}" ({cols_sql}) VALUES ({placeholders})',
                    values,
                )
        conn.commit()

    insert_job(
        job_id=job_id,
        user_id=user_id,
        target_table=T_TAB_MUNICIPAL,
        staging_table=stg,
        status="ready",
        report=report,
    )
    return get_job(job_id) or {"id": job_id, "status": "ready", "report": report}


def _run_etl_007() -> Dict[str, Any]:
    """Post-carga: total_unidades_medicas + backfill habxpol (sql/007)."""
    notes: List[str] = []
    with get_db() as conn:
        with conn.cursor() as cur:
            try:
                cur.execute(
                    f"""
                    ALTER TABLE "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                      ADD COLUMN IF NOT EXISTS total_unidades_medicas double precision
                    """
                )
                cur.execute(
                    f"""
                    UPDATE "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                       SET total_unidades_medicas =
                             COALESCE(imss, 0) + COALESCE(issste, 0) + COALESCE(semar, 0)
                           + COALESCE(imb, 0) + COALESCE(sesa, 0) + COALESCE(ssa, 0)
                    """
                )
                notes.append("total_unidades_medicas actualizado")
            except Exception as exc:
                notes.append(f"total_unidades_medicas omitido: {exc}")
            try:
                # pop_tot es el nombre real frecuente
                pop_col = None
                for cand in ("pop_tot", "pob_tot"):
                    cur.execute(
                        """
                        SELECT 1 FROM information_schema.columns
                         WHERE table_schema=%s AND table_name=%s AND column_name=%s
                        """,
                        (SCHEMA, T_TAB_MUNICIPAL, cand),
                    )
                    if cur.fetchone():
                        pop_col = cand
                        break
                if pop_col:
                    cur.execute(
                        f"""
                        UPDATE "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                           SET habxpol = ("{pop_col}"::numeric / NULLIF(pol_prev::numeric, 0))
                         WHERE habxpol IS NULL
                           AND pol_prev IS NOT NULL
                           AND pol_prev::numeric > 0
                           AND "{pop_col}" IS NOT NULL
                        """
                    )
                    notes.append("habxpol backfill OK")
            except Exception as exc:
                notes.append(f"habxpol omitido: {exc}")
        conn.commit()
    clear_column_cache(SCHEMA, T_TAB_MUNICIPAL)
    return {"ok": True, "notes": notes}


def apply_indicator_job(job_id: str, *, user_id: Optional[int] = None) -> Dict[str, Any]:
    import time

    jid = assert_job_id(job_id)
    job = get_job(jid)
    if not job:
        raise ValueError("JOB_NOT_FOUND")
    if job["status"] != "ready":
        raise ValueError(f"JOB_NOT_READY:{job['status']}")
    if (job.get("report") or {}).get("kind") != "indicator":
        raise ValueError("JOB_NOT_INDICATOR")

    report = dict(job.get("report") or {})
    tgt = str(job.get("target_table") or report.get("target_table") or "").lower()
    if tgt == T_TAB_NACIONAL or report.get("scope") == "nacional":
        from data_refresh.indicator_nacional import apply_nacional_job

        return apply_nacional_job(job_id, user_id=user_id)

    if (report.get("validation") or {}).get("level") == "block":
        raise ValueError("JOB_BLOCKED_VALIDATION")

    stg = (job.get("staging_table") or indicator_staging_name(jid)).strip()
    metrics = report.get("mapped_metrics") or []
    if not metrics:
        raise ValueError("NO_METRICS")

    t0 = time.monotonic()
    update_job(jid, status="applying")

    version_meta = None
    try:
        from data_refresh.versions import snapshot_table_as_version

        version_meta = snapshot_table_as_version(
            table_name=T_TAB_MUNICIPAL, job_id=jid, kind="indicator"
        )
    except Exception as snap_exc:
        logger.warning("Snapshot tab_municipal falló: %s", snap_exc)

    set_parts = []
    for m in metrics:
        col = _qident(m["resolved_column"])
        set_parts.append(f'"{col}" = s."{col}"')
    set_sql = ", ".join(set_parts)

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                # Merge municipal por cve — excluye filas Nacional/Estatal (cve espuria 012, etc.)
                cur.execute(
                    f"""
                    UPDATE "{SCHEMA}"."{T_TAB_MUNICIPAL}" AS t
                       SET {set_sql}
                      FROM "{STAGING_SCHEMA}"."{stg}" AS s
                     WHERE TRIM(BOTH FROM t.cve_mun::text) = TRIM(BOTH FROM s.cve_mun::text)
                       AND TRIM(BOTH FROM s.cve_mun::text) ~ '^[0-9]{{3}}$'
                       AND LOWER(TRIM(BOTH FROM COALESCE(s.nom_mun::text, '')))
                           NOT IN ('nacional', 'estatal')
                       AND LOWER(TRIM(BOTH FROM COALESCE(t.nom_mun::text, '')))
                           NOT IN ('nacional', 'estatal')
                    """
                )
                mun_updated = cur.rowcount
                # Merge especial SOLO por nom_mun (nunca por cve_mun)
                cur.execute(
                    f"""
                    UPDATE "{SCHEMA}"."{T_TAB_MUNICIPAL}" AS t
                       SET {set_sql}
                      FROM "{STAGING_SCHEMA}"."{stg}" AS s
                     WHERE LOWER(TRIM(BOTH FROM t.nom_mun::text))
                         = LOWER(TRIM(BOTH FROM s.nom_mun::text))
                       AND LOWER(TRIM(BOTH FROM s.nom_mun::text)) IN ('nacional', 'estatal')
                       AND LOWER(TRIM(BOTH FROM t.nom_mun::text)) IN ('nacional', 'estatal')
                    """
                )
                special_updated = cur.rowcount
            conn.commit()

        etl = _run_etl_007()
        validate = {}
        try:
            from indicators_service import validate_indicators_catalog

            validate = validate_indicators_catalog()
        except Exception as exc:
            validate = {"ok": False, "error": str(exc)[:200]}

        elapsed = round(time.monotonic() - t0, 2)
        report["applied"] = True
        report["phase"] = "applied"
        report["progress"] = 100
        report["label"] = "Indicadores actualizados"
        report["apply_summary"] = {
            "rows_municipal_updated": mun_updated,
            "rows_special_updated": special_updated,
            "columns_updated": [m["column"] for m in metrics],
            "backup_table": (version_meta or {}).get("backup_table"),
            "version_id": (version_meta or {}).get("id"),
            "etl": etl,
            "validate_ok": bool(validate.get("ok") or validate.get("etl_ready")),
            "validate": {
                "ok": validate.get("ok"),
                "etl_ready": validate.get("etl_ready"),
                "summary": validate.get("summary"),
            },
            "elapsed_seconds": elapsed,
            "checks": [
                {"ok": True, "label": "Merge por columnas aplicado"},
                {"ok": True, "label": f"Municipios actualizados: {mun_updated}"},
                {
                    "ok": True,
                    "label": f"Filas Nacional/Estatal actualizadas: {special_updated}",
                },
                {
                    "ok": bool((version_meta or {}).get("backup_table")),
                    "label": (
                        f"Snapshot: {(version_meta or {}).get('backup_table')}"
                        if (version_meta or {}).get("backup_table")
                        else "Snapshot no disponible"
                    ),
                },
                {
                    "ok": bool(etl.get("ok")),
                    "label": "Post-ETL (007) ejecutado",
                },
                {
                    "ok": bool(validate.get("ok") or validate.get("etl_ready")),
                    "label": (
                        "Catálogo de indicadores validado"
                        if (validate.get("ok") or validate.get("etl_ready"))
                        else "Validación de indicadores con observaciones"
                    ),
                },
            ],
        }
        update_job(jid, status="applied", report=report, error_message="")
        _drop_staging(stg)
        return get_job(jid) or job
    except Exception as exc:
        update_job(jid, status="ready", error_message=str(exc)[:800])
        raise


def cancel_indicator_job(job_id: str) -> Dict[str, Any]:
    jid = assert_job_id(job_id)
    job = get_job(jid)
    if not job:
        raise ValueError("JOB_NOT_FOUND")
    if job["status"] == "applied":
        raise ValueError("JOB_ALREADY_APPLIED")
    if job["status"] == "applying":
        raise ValueError("JOB_BUSY:applying")
    stg = job.get("staging_table")
    if stg:
        try:
            _drop_staging(stg)
        except Exception:
            pass
    update_job(jid, status="cancelled")
    return get_job(jid) or job


def build_template_mold_csv(template_id: str) -> Tuple[str, str]:
    """CSV molde legacy (separador ,). Preferir build_template_mold_xlsx."""
    template = _get_template(template_id)
    metrics = [m for m in (template.get("metrics") or []) if m.get("in_database")]
    if not metrics:
        raise ValueError("TEMPLATE_SIN_COLUMNAS_EN_BD")

    headers = ["cve_mun", "nom_mun"] + [str(m["column"]) for m in metrics]
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=",", quoting=csv.QUOTE_MINIMAL)
    writer.writerow(headers)

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM cve_mun::text) AS cve,
                       TRIM(BOTH FROM nom_mun::text) AS nom
                  FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                 WHERE TRIM(BOTH FROM cve_mun::text) ~ '^[0-9]{{3}}$'
                   AND LOWER(TRIM(BOTH FROM COALESCE(nom_mun::text, '')))
                       NOT IN ('nacional', 'estatal')
                 ORDER BY 1
                """
            )
            mun_rows = cur.fetchall() or []
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM cve_mun::text) AS cve,
                       TRIM(BOTH FROM nom_mun::text) AS nom
                  FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                 WHERE LOWER(TRIM(BOTH FROM nom_mun::text)) IN ('nacional', 'estatal')
                 ORDER BY CASE LOWER(TRIM(BOTH FROM nom_mun::text))
                            WHEN 'nacional' THEN 0
                            WHEN 'estatal' THEN 1
                            ELSE 2
                          END
                """
            )
            special = cur.fetchall() or []

    empty = [""] * len(metrics)
    mun_cves = {_normalize_cve_mun(r.get("cve")) for r in mun_rows}
    for r in mun_rows:
        cve = _normalize_cve_mun(r.get("cve")) or str(r.get("cve") or "")
        nom = str(r.get("nom") or "")
        writer.writerow([f"\t{cve}" if CVE_RE.match(cve) else cve, nom, *empty])
    for r in special:
        nom = str(r.get("nom") or "")
        raw_cve = str(r.get("cve") or "").strip()
        cve = "" if _normalize_cve_mun(raw_cve) in mun_cves else raw_cve
        writer.writerow([cve, nom, *empty])

    tid = str(template.get("id") or "indicador").replace("/", "_")
    return f"molde_{tid}.csv", buf.getvalue()


def build_template_mold_xlsx(template_id: str) -> Tuple[str, bytes]:
    """Molde Excel: municipal (cve_mun 3 dig) o nacional (ent 2 dig)."""
    template = _get_template(template_id)
    if _template_is_nacional(template):
        from data_refresh.indicator_nacional import build_nacional_mold_xlsx

        return build_nacional_mold_xlsx(template)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("XLSX_NO_DISPONIBLE: openpyxl no está instalado") from exc

    metrics = [m for m in (template.get("metrics") or []) if m.get("in_database")]
    if not metrics:
        raise ValueError("TEMPLATE_SIN_COLUMNAS_EN_BD")

    headers = ["cve_mun", "nom_mun"] + [str(m["column"]) for m in metrics]

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM cve_mun::text) AS cve,
                       TRIM(BOTH FROM nom_mun::text) AS nom
                  FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                 WHERE TRIM(BOTH FROM cve_mun::text) ~ '^[0-9]{{3}}$'
                   AND LOWER(TRIM(BOTH FROM COALESCE(nom_mun::text, '')))
                       NOT IN ('nacional', 'estatal')
                 ORDER BY 1
                """
            )
            mun_rows = cur.fetchall() or []
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM cve_mun::text) AS cve,
                       TRIM(BOTH FROM nom_mun::text) AS nom
                  FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                 WHERE LOWER(TRIM(BOTH FROM nom_mun::text)) IN ('nacional', 'estatal')
                 ORDER BY CASE LOWER(TRIM(BOTH FROM nom_mun::text))
                            WHEN 'nacional' THEN 0
                            WHEN 'estatal' THEN 1
                            ELSE 2
                          END
                """
            )
            special = cur.fetchall() or []

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(1, col_i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    mun_cves = {_normalize_cve_mun(r.get("cve")) for r in mun_rows}
    row_i = 2
    for r in mun_rows:
        cve = _normalize_cve_mun(r.get("cve")) or str(r.get("cve") or "")
        nom = str(r.get("nom") or "")
        cve_cell = ws.cell(row_i, 1, cve)
        cve_cell.number_format = "@"
        ws.cell(row_i, 2, nom)
        for col_i in range(3, len(headers) + 1):
            ws.cell(row_i, col_i, None)
        row_i += 1
    for r in special:
        nom = str(r.get("nom") or "")
        raw_cve = str(r.get("cve") or "").strip()
        cve = "" if _normalize_cve_mun(raw_cve) in mun_cves else raw_cve
        cve_cell = ws.cell(row_i, 1, cve)
        cve_cell.number_format = "@"
        ws.cell(row_i, 2, nom)
        for col_i in range(3, len(headers) + 1):
            ws.cell(row_i, col_i, None)
        row_i += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    for col_i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 16

    instr = wb.create_sheet("Instrucciones", 0)
    lines = [
        f"Molde: {template.get('label') or template.get('id')}",
        "",
        "1. Rellene solo las columnas numéricas en la hoja «Datos».",
        "2. No modifique cve_mun de municipios (3 dígitos: 001, 002, …).",
        "3. Para municipios, el cruce se hace por cve_mun.",
        "4. Filas Nacional / Estatal se actualizan SOLO por nom_mun (no por cve).",
        "   Deje cve_mun vacío en esas filas. En BD, Estatal puede tener cve «12»",
        "   (cve_ent); NO use 012 — colisiona con Ayutla de los Libres.",
        "5. Guarde este .xlsx y súbalo en Data Refresh Studio (también acepta CSV).",
        "6. Si exporta a CSV desde Excel, use UTF-8; el servidor acepta , o ;.",
        "",
        "Columnas de métricas:",
    ]
    for m in metrics:
        lines.append(f"  - {m['column']}: {m.get('label') or m['column']}")
    for i, line in enumerate(lines, start=1):
        instr.cell(i, 1, line)
    instr.column_dimensions["A"].width = 90

    out = io.BytesIO()
    wb.save(out)
    tid = str(template.get("id") or "indicador").replace("/", "_")
    return f"molde_{tid}.xlsx", out.getvalue()


def build_synthetic_indicator_xlsx(
    *,
    template_id: str,
    changes: int = 40,
    seed: int = 1234,
    jitter_pct: float = 0.08,
) -> Tuple[str, bytes, Dict[str, Any]]:
    """Excel de prueba: valores actuales de BD + N celdas perturbadas (seed reproducible).

    No escribe en producción. Sirve para subir al Indicator Refresh y validar el pipeline.
    Returns (filename, xlsx_bytes, meta).
    """
    import random

    template = _get_template(template_id)
    if _template_is_nacional(template):
        from data_refresh.indicator_nacional import build_nacional_synthetic_xlsx

        return build_nacional_synthetic_xlsx(
            template=template,
            changes=changes,
            seed=seed,
            jitter_pct=jitter_pct,
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("XLSX_NO_DISPONIBLE: openpyxl no está instalado") from exc

    metrics = [m for m in (template.get("metrics") or []) if m.get("in_database")]
    if not metrics:
        raise ValueError("TEMPLATE_SIN_COLUMNAS_EN_BD")

    n_changes = max(1, min(int(changes), 5000))
    seed_i = int(seed)
    jitter = max(0.01, min(float(jitter_pct), 0.5))
    rng = random.Random(seed_i)

    cols_sql = ", ".join(
        f'"{_qident(m["resolved_column"])}" AS "{m["column"]}"' for m in metrics
    )
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM cve_mun::text) AS cve,
                       TRIM(BOTH FROM nom_mun::text) AS nom,
                       {cols_sql}
                  FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                 WHERE TRIM(BOTH FROM cve_mun::text) ~ '^[0-9]{{3}}$'
                   AND LOWER(TRIM(BOTH FROM COALESCE(nom_mun::text, '')))
                       NOT IN ('nacional', 'estatal')
                 ORDER BY 1
                """
            )
            mun_rows = cur.fetchall() or []
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM cve_mun::text) AS cve,
                       TRIM(BOTH FROM nom_mun::text) AS nom,
                       {cols_sql}
                  FROM "{SCHEMA}"."{T_TAB_MUNICIPAL}"
                 WHERE LOWER(TRIM(BOTH FROM nom_mun::text)) IN ('nacional', 'estatal')
                 ORDER BY CASE LOWER(TRIM(BOTH FROM nom_mun::text))
                            WHEN 'nacional' THEN 0
                            WHEN 'estatal' THEN 1
                            ELSE 2
                          END
                """
            )
            special = cur.fetchall() or []

    if not mun_rows:
        raise ValueError("SIN_MUNICIPIOS_EN_BD")

    # Matriz editable: solo municipios (no Nacional/Estatal en el muestreo de cambios)
    metric_keys = [str(m["column"]) for m in metrics]
    candidates: List[Tuple[int, int]] = []  # (row_idx, metric_idx)
    for ri, r in enumerate(mun_rows):
        for mi, mk in enumerate(metric_keys):
            val = r.get(mk)
            if val is None:
                continue
            try:
                float(val)
            except (TypeError, ValueError):
                continue
            candidates.append((ri, mi))

    if not candidates:
        raise ValueError("SIN_VALORES_NUMERICOS_PARA_PERTURBAR")

    pick_n = min(n_changes, len(candidates))
    picked = rng.sample(candidates, pick_n)
    change_set = {(ri, mi): True for ri, mi in picked}
    change_log: List[Dict[str, Any]] = []

    def _perturb(old: float, typ: str) -> float:
        factor = 1.0 + rng.uniform(-jitter, jitter)
        new_v = float(old) * factor
        if typ == "integer":
            new_v = float(max(0, int(round(new_v))))
        else:
            new_v = round(new_v, 4)
            if new_v < 0:
                new_v = 0.0
        if new_v == float(old):
            # Forzar diferencia mínima observable
            bump = 1.0 if typ == "integer" else max(0.0001, abs(float(old)) * 0.01)
            new_v = float(old) + bump
            if typ == "integer":
                new_v = float(int(round(new_v)))
        return new_v

    # Aplicar perturbaciones sobre copia de valores
    values_by_row: List[Dict[str, Any]] = []
    for ri, r in enumerate(mun_rows):
        row_vals: Dict[str, Any] = {}
        for mi, m in enumerate(metrics):
            mk = metric_keys[mi]
            old = r.get(mk)
            if (ri, mi) in change_set and old is not None:
                try:
                    old_f = float(old)
                except (TypeError, ValueError):
                    row_vals[mk] = old
                    continue
                new_v = _perturb(old_f, str(m.get("type") or "float"))
                row_vals[mk] = new_v
                change_log.append(
                    {
                        "cve_mun": _normalize_cve_mun(r.get("cve")),
                        "nom_mun": str(r.get("nom") or ""),
                        "column": mk,
                        "before": old_f,
                        "after": new_v,
                    }
                )
            else:
                row_vals[mk] = old
        values_by_row.append(row_vals)

    special_vals: List[Dict[str, Any]] = []
    for r in special:
        special_vals.append({mk: r.get(mk) for mk in metric_keys})

    headers = ["cve_mun", "nom_mun"] + metric_keys
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    changed_fill = PatternFill("solid", fgColor="FFF2CC")
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(1, col_i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_i = 2
    for ri, r in enumerate(mun_rows):
        cve = _normalize_cve_mun(r.get("cve")) or str(r.get("cve") or "")
        cve_cell = ws.cell(row_i, 1, cve)
        cve_cell.number_format = "@"
        ws.cell(row_i, 2, str(r.get("nom") or ""))
        for mi, mk in enumerate(metric_keys):
            cell = ws.cell(row_i, 3 + mi, values_by_row[ri].get(mk))
            if (ri, mi) in change_set:
                cell.fill = changed_fill
        row_i += 1

    for si, r in enumerate(special):
        # cve vacío: merge solo por nom_mun (evita 012 = Ayutla)
        cve_cell = ws.cell(row_i, 1, "")
        cve_cell.number_format = "@"
        ws.cell(row_i, 2, str(r.get("nom") or ""))
        for mi, mk in enumerate(metric_keys):
            ws.cell(row_i, 3 + mi, special_vals[si].get(mk))
        row_i += 1

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 32
    for col_i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14

    esc = wb.create_sheet("Escenario", 0)
    bak = f"atlas.tab_municipal_bak_seed{seed_i}"
    meta_lines = [
        "Dataset sintético reproducible — GroSIG Data Refresh",
        f"template_id: {template.get('id')}",
        f"template: {template.get('label') or template.get('id')}",
        f"seed: {seed_i}",
        f"changes_solicitados: {n_changes}",
        f"changes_aplicados: {len(change_log)}",
        f"jitter_pct: ±{jitter * 100:.1f}%",
        f"candidatos: {len(candidates)}",
        "",
        "Este archivo NO modifica la BD. Súbalo en Indicadores → Subir y validar.",
        "Antes de Aplicar, haga backup:",
        f"  CREATE TABLE {bak} AS TABLE atlas.tab_municipal;",
        "Para restaurar:",
        f"  TRUNCATE atlas.tab_municipal; INSERT INTO atlas.tab_municipal SELECT * FROM {bak};",
        "(Ajuste columnas/constraints según su entorno; preferible pg_dump de la tabla.)",
        "",
        "Celdas en amarillo = valores perturbados.",
    ]
    for i, line in enumerate(meta_lines, start=1):
        esc.cell(i, 1, line)
    esc.column_dimensions["A"].width = 100

    log_ws = wb.create_sheet("Cambios")
    log_headers = ["cve_mun", "nom_mun", "column", "before", "after"]
    for col_i, h in enumerate(log_headers, start=1):
        cell = log_ws.cell(1, col_i, h)
        cell.font = Font(bold=True)
    for i, ch in enumerate(change_log, start=2):
        log_ws.cell(i, 1, ch["cve_mun"]).number_format = "@"
        log_ws.cell(i, 2, ch["nom_mun"])
        log_ws.cell(i, 3, ch["column"])
        log_ws.cell(i, 4, ch["before"])
        log_ws.cell(i, 5, ch["after"])

    out = io.BytesIO()
    wb.save(out)
    tid = str(template.get("id") or "indicador").replace("/", "_")
    filename = f"prueba_{tid}_seed{seed_i}_n{len(change_log)}.xlsx"
    meta = {
        "template_id": template.get("id"),
        "seed": seed_i,
        "changes_requested": n_changes,
        "changes_applied": len(change_log),
        "jitter_pct": jitter,
        "candidates": len(candidates),
        "backup_hint": bak,
        "filename": filename,
    }
    return filename, out.getvalue(), meta


def list_indicator_jobs(limit: int = 15) -> List[Dict[str, Any]]:
    jobs = list_recent_jobs(limit * 3)
    out = []
    for j in jobs:
        full = get_job(j["id"])
        if not full:
            continue
        if (full.get("report") or {}).get("kind") == "indicator":
            out.append(
                {
                    "id": full["id"],
                    "target_table": full.get("target_table"),
                    "status": full.get("status"),
                    "created_at": full.get("created_at"),
                    "updated_at": full.get("updated_at"),
                    "error_message": full.get("error_message"),
                    "template_label": (full.get("report") or {}).get("template_label"),
                }
            )
        if len(out) >= limit:
            break
    return out
