"""Indicator Refresh para atlas.tab_nacional (ent 01–32 + especiales por nombre)."""

from __future__ import annotations

import io
import logging
import random
import re
from typing import Any, Dict, List, Optional, Sequence, Tuple

from column_resolver import clear_column_cache, resolve_column
from data_refresh.indicator_refresh import (
    ENT_RE,
    SPECIAL_NOMS,
    _decode_tabular,
    _drop_staging,
    _find_csv_col,
    _norm_header_map,
    _normalize_ent,
    _parse_number,
    _qident,
    ensure_staging_schema,
    indicator_staging_name,
)
from data_refresh.jobs_store import get_job, insert_job, new_job_id, update_job
from data_refresh.names import STAGING_SCHEMA, assert_job_id
from database import get_db
from tables import SCHEMA, T_TAB_NACIONAL

logger = logging.getLogger(__name__)


def _nat_key_cols(conn) -> Tuple[str, str]:
    col_ent = resolve_column(conn, SCHEMA, T_TAB_NACIONAL, ("ent", "ENT", "cve_ent", "CVE_ENT"))
    col_nom = resolve_column(
        conn, SCHEMA, T_TAB_NACIONAL, ("nom_ent", "NOM_ENT", "nomgeo", "NOMGEO")
    )
    if not col_ent or not col_nom:
        raise ValueError("TAB_NACIONAL_SIN_ENT_NOM: faltan columnas ent/nom_ent")
    return col_ent, col_nom


def _is_nat_special(nom: str, ent: str) -> bool:
    if (nom or "").strip().lower() in SPECIAL_NOMS:
        return True
    return not bool(ENT_RE.match(ent or ""))


def prod_entity_ents(col_ent: str, col_nom: str) -> List[str]:
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM "{col_ent}"::text) AS ent
                  FROM "{SCHEMA}"."{T_TAB_NACIONAL}"
                 WHERE TRIM(BOTH FROM "{col_ent}"::text) ~ '^[0-9]{{2}}$'
                   AND TRIM(BOTH FROM "{col_ent}"::text)::int BETWEEN 1 AND 32
                   AND LOWER(TRIM(BOTH FROM COALESCE("{col_nom}"::text, '')))
                       NOT IN ('nacional', 'estatal')
                 ORDER BY 1
                """
            )
            out = []
            for r in cur.fetchall() or []:
                e = _normalize_ent(r.get("ent"))
                if ENT_RE.match(e):
                    out.append(e)
            return out


def prod_special_noms_nat(col_nom: str) -> Dict[str, bool]:
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT LOWER(TRIM(BOTH FROM "{col_nom}"::text)) AS nom
                  FROM "{SCHEMA}"."{T_TAB_NACIONAL}"
                 WHERE LOWER(TRIM(BOTH FROM "{col_nom}"::text)) IN ('nacional', 'estatal')
                """
            )
            found = {str(r["nom"]) for r in (cur.fetchall() or []) if r.get("nom")}
    return {"nacional": "nacional" in found, "estatal": "estatal" in found}


def build_nacional_template_entry(
    conn, ind: Dict[str, Any], fields: Sequence[Dict[str, Any]]
) -> Optional[Dict[str, Any]]:
    metrics: List[Dict[str, Any]] = []
    for f in fields:
        st = str(f.get("source_table") or "").lower()
        if st != "tab_nacional":
            continue
        col = str(f.get("column") or f.get("key") or "").strip()
        if not col:
            continue
        aliases = [col] + list(f.get("column_aliases") or [])
        resolved = resolve_column(conn, SCHEMA, T_TAB_NACIONAL, aliases)
        metrics.append(
            {
                "key": f.get("key") or col,
                "column": col,
                "resolved_column": resolved,
                "aliases": aliases,
                "label": f.get("label") or col,
                "type": f.get("type") or "float",
                "in_database": bool(resolved),
            }
        )
    if not metrics:
        return None
    headers = ["ent", "nom_ent"] + [m["column"] for m in metrics if m.get("in_database")]
    base_id = str(ind.get("id") or "indicador")
    return {
        "id": f"{base_id}__nacional",
        "indicator_id": base_id,
        "label": ind.get("label") or base_id,
        "group_id": ind.get("group_id"),
        "subtitle": ind.get("subtitle") or "",
        "table": T_TAB_NACIONAL,
        "target_table": T_TAB_NACIONAL,
        "scope": "nacional",
        "key_columns": ["ent", "nom_ent"],
        "metrics": metrics,
        "csv_headers_hint": ";".join(headers),
        "metrics_in_db": sum(1 for m in metrics if m.get("in_database")),
    }


def build_report_nacional(
    *,
    template: Dict[str, Any],
    headers: List[str],
    rows: List[Dict[str, str]],
    mapped_metrics: List[Dict[str, Any]],
    ent_col: str,
    nom_col: Optional[str],
    col_ent_db: str,
    col_nom_db: str,
) -> Dict[str, Any]:
    prod_ents = set(prod_entity_ents(col_ent_db, col_nom_db))
    csv_ents: set = set()
    csv_special = {"nacional": False, "estatal": False}
    bad_ent = 0
    null_counts = {m["resolved_column"]: 0 for m in mapped_metrics}
    type_errors: List[str] = []
    collision_warnings: List[str] = []

    for row in rows:
        nom = (row.get(nom_col) or "").strip().lower() if nom_col else ""
        ent = _normalize_ent(row.get(ent_col))
        is_special = _is_nat_special(nom, ent)
        if is_special:
            if nom in SPECIAL_NOMS:
                csv_special[nom] = True
            elif nom:
                # fila especial por ent inválido
                pass
            if ENT_RE.match(ent) and ent in prod_ents:
                collision_warnings.append(
                    f"Fila especial «{nom or ent}» trae ent={ent} de una entidad real; "
                    "el merge usará solo nom_ent."
                )
        elif ENT_RE.match(ent):
            if ent in csv_ents:
                collision_warnings.append(f"ent duplicada en archivo: {ent}")
            csv_ents.add(ent)
        else:
            bad_ent += 1

        for m in mapped_metrics:
            csv_h = m["csv_header"]
            raw = row.get(csv_h, "")
            if raw == "":
                null_counts[m["resolved_column"]] += 1
                continue
            try:
                _parse_number(raw, m.get("type") or "float")
            except ValueError:
                type_errors.append(
                    f"Fila ent={ent or nom!r} col={m['column']}: valor {raw!r}"
                )

    missing = sorted(prod_ents - csv_ents)
    extra = sorted(csv_ents - prod_ents)
    prod_special = prod_special_noms_nat(col_nom_db)

    warnings: List[str] = list(dict.fromkeys(collision_warnings))
    infos: List[str] = []
    if missing:
        warnings.append(
            f"Faltan {len(missing)} entidades del padrón actual "
            f"(ej. {', '.join(missing[:8])}{'…' if len(missing) > 8 else ''})."
        )
    if extra:
        warnings.append(
            f"Hay {len(extra)} ent no presentes en producción "
            "(no se insertarán filas nuevas en MVP merge)."
        )
    if bad_ent:
        warnings.append(f"{bad_ent} filas con ent inválido (se ignoran en merge por entidad).")
    if type_errors:
        warnings.append(
            f"{len(type_errors)} valores no numéricos (ej. {type_errors[0]})."
        )
    for key, present_prod in prod_special.items():
        if present_prod and not csv_special.get(key):
            infos.append(
                f"Fila «{key.capitalize()}» no viene en el archivo; se conserva la de producción."
            )
        elif csv_special.get(key):
            infos.append(f"Fila «{key.capitalize()}» incluida; se actualizará por nom_ent.")

    level = "ok"
    label = "Listo para aplicar"
    dup_msgs = [w for w in warnings if "duplicada" in w.lower()]
    if type_errors or (missing and len(missing) > 10) or dup_msgs:
        level = "warn"
        label = "Revisar advertencias"
    if not mapped_metrics:
        level = "block"
        label = "Sin columnas métricas válidas"
    if not csv_ents and not any(csv_special.values()):
        level = "block"
        label = "Sin filas de entidad válidas"
    if dup_msgs and len(dup_msgs) > 5:
        level = "block"
        label = "Demasiadas claves duplicadas"

    deltas: List[Dict[str, Any]] = []
    if mapped_metrics and csv_ents:
        m0 = mapped_metrics[0]
        col = m0["resolved_column"]
        csv_h = m0["csv_header"]
        changed = 0
        sample: List[Dict[str, Any]] = []
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    SELECT TRIM(BOTH FROM "{col_ent_db}"::text) AS ent,
                           "{col}" AS val
                      FROM "{SCHEMA}"."{T_TAB_NACIONAL}"
                     WHERE TRIM(BOTH FROM "{col_ent_db}"::text) ~ '^[0-9]{{2}}$'
                       AND TRIM(BOTH FROM "{col_ent_db}"::text)::int BETWEEN 1 AND 32
                       AND LOWER(TRIM(BOTH FROM COALESCE("{col_nom_db}"::text, '')))
                           NOT IN ('nacional', 'estatal')
                    """
                )
                prod_map = {}
                for r in cur.fetchall() or []:
                    e = _normalize_ent(r.get("ent"))
                    if e:
                        prod_map[e] = r.get("val")
        for row in rows:
            nom = (row.get(nom_col) or "").strip().lower() if nom_col else ""
            ent = _normalize_ent(row.get(ent_col))
            if _is_nat_special(nom, ent):
                continue
            if ent not in prod_ents:
                continue
            raw = row.get(csv_h, "")
            try:
                new_v = _parse_number(raw, m0.get("type") or "float")
            except ValueError:
                continue
            old_v = prod_map.get(ent)
            try:
                old_f = float(old_v) if old_v is not None else None
            except (TypeError, ValueError):
                old_f = None
            if old_f != new_v:
                changed += 1
                if len(sample) < 5:
                    sample.append(
                        {
                            "ent": ent,
                            "before": old_f,
                            "after": new_v,
                            "column": m0["column"],
                        }
                    )
        deltas.append(
            {
                "column": m0["column"],
                "label": m0.get("label"),
                "rows_changed": changed,
                "sample": sample,
            }
        )

    ok_n = len(csv_ents & prod_ents)
    return {
        "kind": "indicator",
        "strategy": "column_merge_by_ent",
        "strategy_label": "Actualización por columnas (merge nacional)",
        "template_id": template.get("id"),
        "template_label": template.get("label"),
        "target_table": T_TAB_NACIONAL,
        "scope": "nacional",
        "csv_headers": headers,
        "mapped_metrics": mapped_metrics,
        "row_count_csv": len(rows),
        "municipal_in_csv": len(csv_ents),
        "municipal_matched": ok_n,
        "municipal_missing": missing,
        "municipal_extra": extra,
        "entities_in_csv": len(csv_ents),
        "entities_matched": ok_n,
        "entities_missing": missing,
        "entities_extra": extra,
        "special_rows_csv": csv_special,
        "null_counts": null_counts,
        "type_error_count": len(type_errors),
        "type_error_samples": type_errors[:8],
        "deltas": deltas,
        "warnings": warnings,
        "infos": infos
        + [
            "MVP: solo se actualizan columnas del archivo; el resto de tab_nacional se conserva.",
            "Entidades: merge por ent (2 dígitos). Filas Nacional/Estatal: solo por nom_ent.",
        ],
        "validation": {
            "level": level,
            "label": label,
            "checks": {
                "template": True,
                "metrics": bool(mapped_metrics),
                "municipalities": ok_n > 0 or any(csv_special.values()),
                "types": len(type_errors) == 0,
                "duplicates": len(dup_msgs) == 0,
                "swap_available": level != "block",
            },
            "checklist": [
                {
                    "id": "metrics",
                    "label": "Columnas / métricas mapeadas",
                    "status": "ok" if mapped_metrics else "block",
                    "detail": f"{len(mapped_metrics)} cols",
                },
                {
                    "id": "entities",
                    "label": "Claves ent válidas",
                    "status": "ok" if ok_n > 0 or any(csv_special.values()) else "block",
                    "detail": f"{ok_n} matched",
                },
                {
                    "id": "duplicates",
                    "label": "Duplicados de clave",
                    "status": "warn" if dup_msgs else "ok",
                    "detail": f"{len(dup_msgs)} avisos" if dup_msgs else "OK",
                },
                {
                    "id": "types",
                    "label": "Tipos numéricos",
                    "status": "warn" if type_errors else "ok",
                    "detail": f"{len(type_errors)} errores",
                },
            ],
        },
        "summary_line": (
            f"{ok_n} entidades OK"
            + (f", {len(missing)} faltantes" if missing else ", 0 faltantes")
        ),
        "phase": "ready",
        "progress": 100,
        "label": "Validación lista",
    }


def create_nacional_job(
    *,
    content: bytes,
    filename: str,
    template: Dict[str, Any],
    user_id: Optional[int] = None,
) -> Dict[str, Any]:
    with get_db() as conn:
        col_ent_db, col_nom_db = _nat_key_cols(conn)

    headers, rows = _decode_tabular(content, filename)
    header_map = _norm_header_map(headers)

    ent_col = _find_csv_col(header_map, ("ent", "cve_ent", "CVE_ENT", "ENT"))
    if not ent_col:
        raise ValueError("CSV_FALTA_ENT: se requiere columna ent (o cve_ent)")
    nom_col = _find_csv_col(
        header_map, ("nom_ent", "nomgeo", "nombre", "NOM_ENT", "NOMGEO")
    )

    mapped_metrics: List[Dict[str, Any]] = []
    for m in template.get("metrics") or []:
        if not m.get("in_database") or not m.get("resolved_column"):
            continue
        csv_h = _find_csv_col(
            header_map,
            [m["column"], m.get("key") or "", *(m.get("aliases") or [])],
        )
        if not csv_h:
            continue
        mapped_metrics.append({**m, "csv_header": csv_h})

    if not mapped_metrics:
        raise ValueError(
            "CSV_SIN_METRICAS: ninguna columna de la plantilla está en el archivo "
            f"(esperadas: {template.get('csv_headers_hint')})"
        )

    report = build_report_nacional(
        template=template,
        headers=headers,
        rows=rows,
        mapped_metrics=mapped_metrics,
        ent_col=ent_col,
        nom_col=nom_col,
        col_ent_db=col_ent_db,
        col_nom_db=col_nom_db,
    )
    report["filename"] = filename or "upload.xlsx"
    report["nat_key_cols"] = {"ent": col_ent_db, "nom_ent": col_nom_db}

    if report["validation"]["level"] == "block":
        job_id = new_job_id()
        stg = indicator_staging_name(job_id)
        insert_job(
            job_id=job_id,
            user_id=user_id,
            target_table=T_TAB_NACIONAL,
            staging_table=stg,
            status="failed",
            report=report,
            error_message=report["validation"]["label"],
        )
        raise ValueError(report["validation"]["label"])

    ensure_staging_schema()
    job_id = new_job_id()
    stg = indicator_staging_name(job_id)
    _drop_staging(stg)

    col_defs = ['"ent" text', '"nom_ent" text']
    metric_cols = []
    for m in mapped_metrics:
        rc = _qident(m["resolved_column"])
        metric_cols.append(rc)
        col_defs.append(f'"{rc}" double precision')

    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f'CREATE TABLE "{STAGING_SCHEMA}"."{stg}" ({", ".join(col_defs)})'
            )
            for row in rows:
                nom = (row.get(nom_col) or "").strip() if nom_col else ""
                ent = _normalize_ent(row.get(ent_col))
                if _is_nat_special(nom, ent):
                    if nom.lower() not in SPECIAL_NOMS:
                        continue
                    ent = ""
                elif not ENT_RE.match(ent):
                    continue
                values: List[Any] = [ent or None, nom or None]
                for m in mapped_metrics:
                    raw = row.get(m["csv_header"], "")
                    try:
                        values.append(_parse_number(raw, m.get("type") or "float"))
                    except ValueError:
                        values.append(None)
                placeholders = ", ".join(["%s"] * len(values))
                cols_sql = ", ".join(
                    ['"ent"', '"nom_ent"'] + [f'"{c}"' for c in metric_cols]
                )
                cur.execute(
                    f'INSERT INTO "{STAGING_SCHEMA}"."{stg}" ({cols_sql}) VALUES ({placeholders})',
                    values,
                )
        conn.commit()

    insert_job(
        job_id=job_id,
        user_id=user_id,
        target_table=T_TAB_NACIONAL,
        staging_table=stg,
        status="ready",
        report=report,
    )
    return get_job(job_id) or {"id": job_id, "status": "ready", "report": report}


def apply_nacional_job(job_id: str, *, user_id: Optional[int] = None) -> Dict[str, Any]:
    import time

    jid = assert_job_id(job_id)
    job = get_job(jid)
    if not job:
        raise ValueError("JOB_NOT_FOUND")
    if job["status"] != "ready":
        raise ValueError(f"JOB_NOT_READY:{job['status']}")
    report = dict(job.get("report") or {})
    if report.get("kind") != "indicator":
        raise ValueError("JOB_NOT_INDICATOR")
    if (report.get("validation") or {}).get("level") == "block":
        raise ValueError("JOB_BLOCKED_VALIDATION")

    stg = (job.get("staging_table") or indicator_staging_name(jid)).strip()
    metrics = report.get("mapped_metrics") or []
    if not metrics:
        raise ValueError("NO_METRICS")

    keys = report.get("nat_key_cols") or {}
    col_ent_db = keys.get("ent")
    col_nom_db = keys.get("nom_ent")
    if not col_ent_db or not col_nom_db:
        with get_db() as conn:
            col_ent_db, col_nom_db = _nat_key_cols(conn)

    col_ent_db = _qident(col_ent_db)
    col_nom_db = _qident(col_nom_db)

    t0 = time.monotonic()
    update_job(jid, status="applying")

    version_meta = None
    try:
        from data_refresh.versions import snapshot_table_as_version

        version_meta = snapshot_table_as_version(
            table_name=T_TAB_NACIONAL, job_id=jid, kind="indicator"
        )
    except Exception as snap_exc:
        logger.warning("Snapshot tab_nacional falló: %s", snap_exc)

    set_parts = []
    for m in metrics:
        col = _qident(m["resolved_column"])
        # PostgreSQL no admite alias en SET (t."col" → error column "t")
        set_parts.append(f'"{col}" = s."{col}"')
    set_sql = ", ".join(set_parts)

    try:
        with get_db() as conn:
            with conn.cursor() as cur:
                cur.execute(
                    f"""
                    UPDATE "{SCHEMA}"."{T_TAB_NACIONAL}" AS t
                       SET {set_sql}
                      FROM "{STAGING_SCHEMA}"."{stg}" AS s
                     WHERE TRIM(BOTH FROM t."{col_ent_db}"::text)
                         = TRIM(BOTH FROM s.ent::text)
                       AND TRIM(BOTH FROM s.ent::text) ~ '^[0-9]{{2}}$'
                       AND TRIM(BOTH FROM s.ent::text)::int BETWEEN 1 AND 32
                       AND LOWER(TRIM(BOTH FROM COALESCE(s.nom_ent::text, '')))
                           NOT IN ('nacional', 'estatal')
                       AND LOWER(TRIM(BOTH FROM COALESCE(t."{col_nom_db}"::text, '')))
                           NOT IN ('nacional', 'estatal')
                    """
                )
                ent_updated = cur.rowcount
                cur.execute(
                    f"""
                    UPDATE "{SCHEMA}"."{T_TAB_NACIONAL}" AS t
                       SET {set_sql}
                      FROM "{STAGING_SCHEMA}"."{stg}" AS s
                     WHERE LOWER(TRIM(BOTH FROM t."{col_nom_db}"::text))
                         = LOWER(TRIM(BOTH FROM s.nom_ent::text))
                       AND LOWER(TRIM(BOTH FROM s.nom_ent::text)) IN ('nacional', 'estatal')
                       AND LOWER(TRIM(BOTH FROM t."{col_nom_db}"::text)) IN ('nacional', 'estatal')
                    """
                )
                special_updated = cur.rowcount
            conn.commit()

        clear_column_cache(SCHEMA, T_TAB_NACIONAL)
        validate: Dict[str, Any] = {}
        try:
            from indicators_service import validate_indicators_catalog

            validate = validate_indicators_catalog()
        except Exception as exc:
            validate = {"ok": False, "error": str(exc)[:200]}

        elapsed = round(time.monotonic() - t0, 2)
        report["applied"] = True
        report["phase"] = "applied"
        report["progress"] = 100
        report["label"] = "Indicadores nacionales actualizados"
        report["apply_summary"] = {
            "rows_entities_updated": ent_updated,
            "rows_municipal_updated": ent_updated,
            "rows_special_updated": special_updated,
            "columns_updated": [m["column"] for m in metrics],
            "backup_table": (version_meta or {}).get("backup_table"),
            "version_id": (version_meta or {}).get("id"),
            "etl": {"ok": True, "notes": ["ETL 007 no aplica a tab_nacional"]},
            "validate_ok": bool(validate.get("ok") or validate.get("etl_ready")),
            "validate": {
                "ok": validate.get("ok"),
                "etl_ready": validate.get("etl_ready"),
                "summary": validate.get("summary"),
            },
            "elapsed_seconds": elapsed,
            "checks": [
                {"ok": True, "label": "Merge por columnas aplicado (tab_nacional)"},
                {"ok": True, "label": f"Entidades actualizadas: {ent_updated}"},
                {
                    "ok": True,
                    "label": f"Filas Nacional/Estatal actualizadas: {special_updated}",
                },
                {
                    "ok": bool((version_meta or {}).get("backup_table")),
                    "label": (
                        f"Snapshot: {(version_meta or {}).get('backup_table')}"
                        if (version_meta or {}).get("backup_table")
                        else "Snapshot no disponible"
                    ),
                },
                {
                    "ok": bool(validate.get("ok") or validate.get("etl_ready")),
                    "label": (
                        "Catálogo de indicadores validado"
                        if (validate.get("ok") or validate.get("etl_ready"))
                        else "Validación de indicadores con observaciones"
                    ),
                },
            ],
        }
        update_job(jid, status="applied", report=report, error_message="")
        _drop_staging(stg)
        return get_job(jid) or job
    except Exception as exc:
        update_job(jid, status="ready", error_message=str(exc)[:800])
        raise


def _fetch_nat_rows(
    metrics: List[Dict[str, Any]], col_ent: str, col_nom: str
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    cols_sql = ", ".join(
        f'"{_qident(m["resolved_column"])}" AS "{m["column"]}"' for m in metrics
    )
    with get_db() as conn:
        with conn.cursor() as cur:
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM "{col_ent}"::text) AS ent,
                       TRIM(BOTH FROM "{col_nom}"::text) AS nom,
                       {cols_sql}
                  FROM "{SCHEMA}"."{T_TAB_NACIONAL}"
                 WHERE TRIM(BOTH FROM "{col_ent}"::text) ~ '^[0-9]{{2}}$'
                   AND TRIM(BOTH FROM "{col_ent}"::text)::int BETWEEN 1 AND 32
                   AND LOWER(TRIM(BOTH FROM COALESCE("{col_nom}"::text, '')))
                       NOT IN ('nacional', 'estatal')
                 ORDER BY 1
                """
            )
            ents = cur.fetchall() or []
            cur.execute(
                f"""
                SELECT TRIM(BOTH FROM "{col_ent}"::text) AS ent,
                       TRIM(BOTH FROM "{col_nom}"::text) AS nom,
                       {cols_sql}
                  FROM "{SCHEMA}"."{T_TAB_NACIONAL}"
                 WHERE LOWER(TRIM(BOTH FROM "{col_nom}"::text)) IN ('nacional', 'estatal')
                 ORDER BY CASE LOWER(TRIM(BOTH FROM "{col_nom}"::text))
                            WHEN 'nacional' THEN 0
                            WHEN 'estatal' THEN 1
                            ELSE 2
                          END
                """
            )
            special = cur.fetchall() or []
    return ents, special


def build_nacional_mold_xlsx(template: Dict[str, Any]) -> Tuple[str, bytes]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("XLSX_NO_DISPONIBLE: openpyxl no está instalado") from exc

    metrics = [m for m in (template.get("metrics") or []) if m.get("in_database")]
    if not metrics:
        raise ValueError("TEMPLATE_SIN_COLUMNAS_EN_BD")

    with get_db() as conn:
        col_ent, col_nom = _nat_key_cols(conn)
    ents, special = _fetch_nat_rows(metrics, col_ent, col_nom)
    # mold: only keys, empty metrics — strip metric values
    headers = ["ent", "nom_ent"] + [str(m["column"]) for m in metrics]

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(1, col_i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_i = 2
    for r in ents:
        ent = _normalize_ent(r.get("ent"))
        cell = ws.cell(row_i, 1, ent)
        cell.number_format = "@"
        ws.cell(row_i, 2, str(r.get("nom") or ""))
        for col_i in range(3, len(headers) + 1):
            ws.cell(row_i, col_i, None)
        row_i += 1
    for r in special:
        cell = ws.cell(row_i, 1, "")
        cell.number_format = "@"
        ws.cell(row_i, 2, str(r.get("nom") or ""))
        for col_i in range(3, len(headers) + 1):
            ws.cell(row_i, col_i, None)
        row_i += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    for col_i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14

    instr = wb.create_sheet("Instrucciones", 0)
    lines = [
        f"Molde nacional: {template.get('label') or template.get('id')}",
        "",
        "1. Rellene solo las columnas numéricas en la hoja «Datos».",
        "2. No modifique ent de entidades (2 dígitos: 01, 02, …, 32).",
        "3. Para entidades, el cruce se hace por ent (nunca padear a 3 dígitos).",
        "4. Filas Nacional / Estatal se actualizan SOLO por nom_ent (deje ent vacío).",
        "5. Guarde este .xlsx y súbalo en Data Refresh Studio.",
        "",
        "Columnas de métricas:",
    ]
    for m in metrics:
        lines.append(f"  - {m['column']}: {m.get('label') or m['column']}")
    for i, line in enumerate(lines, start=1):
        instr.cell(i, 1, line)
    instr.column_dimensions["A"].width = 90

    out = io.BytesIO()
    wb.save(out)
    tid = str(template.get("id") or "indicador").replace("/", "_")
    return f"molde_{tid}.xlsx", out.getvalue()


def build_nacional_synthetic_xlsx(
    *,
    template: Dict[str, Any],
    changes: int = 40,
    seed: int = 1234,
    jitter_pct: float = 0.08,
) -> Tuple[str, bytes, Dict[str, Any]]:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("XLSX_NO_DISPONIBLE: openpyxl no está instalado") from exc

    metrics = [m for m in (template.get("metrics") or []) if m.get("in_database")]
    if not metrics:
        raise ValueError("TEMPLATE_SIN_COLUMNAS_EN_BD")

    n_changes = max(1, min(int(changes), 5000))
    seed_i = int(seed)
    jitter = max(0.01, min(float(jitter_pct), 0.5))
    rng = random.Random(seed_i)

    with get_db() as conn:
        col_ent, col_nom = _nat_key_cols(conn)
    ents, special = _fetch_nat_rows(metrics, col_ent, col_nom)
    if not ents:
        raise ValueError("SIN_ENTIDADES_EN_BD")

    metric_keys = [str(m["column"]) for m in metrics]
    candidates: List[Tuple[int, int]] = []
    for ri, r in enumerate(ents):
        for mi, mk in enumerate(metric_keys):
            val = r.get(mk)
            if val is None:
                continue
            try:
                float(val)
            except (TypeError, ValueError):
                continue
            candidates.append((ri, mi))
    if not candidates:
        raise ValueError("SIN_VALORES_NUMERICOS_PARA_PERTURBAR")

    pick_n = min(n_changes, len(candidates))
    picked = rng.sample(candidates, pick_n)
    change_set = {(ri, mi): True for ri, mi in picked}
    change_log: List[Dict[str, Any]] = []

    def _perturb(old: float, typ: str) -> float:
        factor = 1.0 + rng.uniform(-jitter, jitter)
        new_v = float(old) * factor
        if typ == "integer":
            new_v = float(max(0, int(round(new_v))))
        else:
            new_v = round(new_v, 4)
            if new_v < 0:
                new_v = 0.0
        if new_v == float(old):
            bump = 1.0 if typ == "integer" else max(0.0001, abs(float(old)) * 0.01)
            new_v = float(old) + bump
            if typ == "integer":
                new_v = float(int(round(new_v)))
        return new_v

    values_by_row: List[Dict[str, Any]] = []
    for ri, r in enumerate(ents):
        row_vals: Dict[str, Any] = {}
        for mi, m in enumerate(metrics):
            mk = metric_keys[mi]
            old = r.get(mk)
            if (ri, mi) in change_set and old is not None:
                try:
                    old_f = float(old)
                except (TypeError, ValueError):
                    row_vals[mk] = old
                    continue
                new_v = _perturb(old_f, str(m.get("type") or "float"))
                row_vals[mk] = new_v
                change_log.append(
                    {
                        "ent": _normalize_ent(r.get("ent")),
                        "nom_ent": str(r.get("nom") or ""),
                        "column": mk,
                        "before": old_f,
                        "after": new_v,
                    }
                )
            else:
                row_vals[mk] = old
        values_by_row.append(row_vals)

    special_vals = [{mk: r.get(mk) for mk in metric_keys} for r in special]
    headers = ["ent", "nom_ent"] + metric_keys

    wb = Workbook()
    ws = wb.active
    ws.title = "Datos"
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    changed_fill = PatternFill("solid", fgColor="FFF2CC")
    for col_i, h in enumerate(headers, start=1):
        cell = ws.cell(1, col_i, h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    row_i = 2
    for ri, r in enumerate(ents):
        ent = _normalize_ent(r.get("ent"))
        cell = ws.cell(row_i, 1, ent)
        cell.number_format = "@"
        ws.cell(row_i, 2, str(r.get("nom") or ""))
        for mi, mk in enumerate(metric_keys):
            cell = ws.cell(row_i, 3 + mi, values_by_row[ri].get(mk))
            if (ri, mi) in change_set:
                cell.fill = changed_fill
        row_i += 1
    for si, r in enumerate(special):
        cell = ws.cell(row_i, 1, "")
        cell.number_format = "@"
        ws.cell(row_i, 2, str(r.get("nom") or ""))
        for mi, mk in enumerate(metric_keys):
            ws.cell(row_i, 3 + mi, special_vals[si].get(mk))
        row_i += 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 28
    for col_i in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = 14

    esc = wb.create_sheet("Escenario", 0)
    bak = f"atlas.tab_nacional_bak_seed{seed_i}"
    meta_lines = [
        "Dataset sintético nacional — GroSIG Data Refresh",
        f"template_id: {template.get('id')}",
        f"seed: {seed_i}",
        f"changes_aplicados: {len(change_log)}",
        f"jitter_pct: ±{jitter * 100:.1f}%",
        "",
        "No escribe en producción. Antes de Aplicar:",
        f"  CREATE TABLE {bak} AS TABLE atlas.tab_nacional;",
        "Restaurar:",
        f"  TRUNCATE atlas.tab_nacional; INSERT INTO atlas.tab_nacional SELECT * FROM {bak};",
    ]
    for i, line in enumerate(meta_lines, start=1):
        esc.cell(i, 1, line)
    esc.column_dimensions["A"].width = 100

    log_ws = wb.create_sheet("Cambios")
    for col_i, h in enumerate(["ent", "nom_ent", "column", "before", "after"], start=1):
        log_ws.cell(1, col_i, h).font = Font(bold=True)
    for i, ch in enumerate(change_log, start=2):
        log_ws.cell(i, 1, ch["ent"]).number_format = "@"
        log_ws.cell(i, 2, ch["nom_ent"])
        log_ws.cell(i, 3, ch["column"])
        log_ws.cell(i, 4, ch["before"])
        log_ws.cell(i, 5, ch["after"])

    out = io.BytesIO()
    wb.save(out)
    tid = str(template.get("id") or "indicador").replace("/", "_")
    filename = f"prueba_{tid}_seed{seed_i}_n{len(change_log)}.xlsx"
    meta = {
        "template_id": template.get("id"),
        "seed": seed_i,
        "changes_requested": n_changes,
        "changes_applied": len(change_log),
        "jitter_pct": jitter,
        "candidates": len(candidates),
        "backup_hint": bak,
        "filename": filename,
        "target_table": T_TAB_NACIONAL,
    }
    return filename, out.getvalue(), meta
