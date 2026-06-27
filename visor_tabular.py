"""
Consulta tabular de capas del visor geográfico.

Expone localidades (c_loc_punto), establecimientos de salud (c_clues) y capas DENUE
filtradas por municipio (cve_mun).
"""

from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Mapping, Optional, Sequence, Tuple

from column_resolver import resolve_column
from tables import SCHEMA, T_CLUES, T_DENUE, T_LOC_PUNTO, qualified
from utils import mun_where_sql, norm_cve_mun, quote_ident
from visor_catalog_loader import load_visor_catalog_raw, ordered_layer_ids_from_raw
from visor_layers import denue_codigos_for_layer, layer_config

logger = logging.getLogger(__name__)

MAX_TABULAR_ROWS = 25_000

NUMERO_FIELD = "num"
NUMERO_LABEL = "No."

TABULAR_ERRORS = {
    "UNKNOWN_LAYER": "Capa no disponible para consulta tabular.",
    "MISSING_CVE_MUN": "Selecciona un municipio en el explorador.",
    "NO_ROWS": "No hay registros para este municipio en la capa seleccionada.",
    "EXPORT_FAILED": "No se pudo generar el archivo Excel.",
}

# (clave_respuesta, candidatos_en_bd, etiqueta_columna)
_LOCSPUNTO_FIELD_SPECS: Sequence[Tuple[str, Sequence[str], str]] = (
    ("cvegeo", ("cvegeo", "CVEGEO"), "Clave geográfica"),
    ("cve_ent", ("cve_ent", "CVE_ENT"), "Clave de entidad"),
    ("nom_ent", ("nom_ent", "NOM_ENT"), "Nombre de entidad"),
    ("cve_mun", ("cve_mun", "CVE_MUN"), "Clave del municipio"),
    ("nom_mun", ("nom_mun", "NOM_MUN"), "Nombre del municipio"),
    ("cve_loc", ("cve_loc", "CVE_LOC"), "Clave de la localidad"),
    ("nom_loc", ("nom_loc", "NOM_LOC"), "Nombre de la localidad"),
    ("ambito", ("ambito", "AMBITO"), "Ámbito"),
    ("altitud", ("altitud", "ALTITUD"), "Altitud"),
    ("pob_total", ("pob_total", "POB_TOTAL"), "Población total"),
    ("pob_mascul", ("pob_mascul", "POB_MASCUL"), "Población masculina"),
    ("pob_femeni", ("pob_femeni", "POB_FEMENI"), "Población femenina"),
    (
        "total_viv",
        (
            "total de v",
            "total_de_v",
            "total_de_viv",
            "total_viv",
            "vivtot",
            "total_viviendas",
        ),
        "Total de viviendas",
    ),
)

_CLUES_FIELD_SPECS: Sequence[Tuple[str, Sequence[str], str]] = (
    ("nom_insti", ("nom_insti",), "Nombre de la institución"),
    ("cve_mun", ("cve_mun",), "Clave del municipio"),
    ("mun", ("mun", "nom_mun", "municipio"), "Nombre del municipio"),
    ("cve_loc", ("cve_loc",), "Clave de la localidad"),
    ("loc", ("loc", "nom_loc", "localidad"), "Nombre de la localidad"),
    ("nom_comer", ("nom_comer",), "Nombre comercial"),
    ("nom_insadm", ("nom_insadm",), "Nombre de la institución administradora"),
)

_DENUE_FIELD_SPECS: Sequence[Tuple[str, Sequence[str], str]] = (
    ("nom_estab", ("nom_estab",), "Nombre del establecimiento"),
    ("nombre_act", ("nombre_act",), "Nombre de la actividad"),
    ("cve_mun", ("cve_mun",), "Clave del municipio"),
    ("municipio", ("municipio", "mun"), "Nombre del municipio"),
    ("cve_loc", ("cve_loc",), "Clave de la localidad"),
    ("localidad", ("localidad", "loc"), "Nombre de la localidad"),
)

_DENUE_DOMICILIO_PARTS: Sequence[Sequence[str]] = (
    ("tipo_vial",),
    ("nom_vial",),
    ("numero_ext", "num_ext"),
    ("tipo_asent",),
    ("nomb_asent", "nom_asent"),
    ("cod_postal",),
)

_DENUE_TABULAR_LAYER_IDS: Sequence[str] = (
    "denue_rastros",
    "denue_gasolinerias",
    "denue_gaseras",
    "denue_escuelas",
    "denue_hospitales",
    "denue_museos",
    "denue_cementerios",
    "denue_iglesias",
)

_TABULAR_PRESETS = frozenset({"locspunto", "clues", "denue"})


def _raw_layer_entry(layer_id: str) -> Optional[Dict[str, Any]]:
    key = (layer_id or "").strip().lower()
    layers = load_visor_catalog_raw().get("layers") or {}
    entry = layers.get(key)
    return entry if isinstance(entry, dict) else None


def _tabular_block(layer_id: str) -> Dict[str, Any]:
    entry = _raw_layer_entry(layer_id) or {}
    block = entry.get("tabular")
    return block if isinstance(block, dict) else {}


def _tabular_preset(layer_id: str) -> Optional[str]:
    block = _tabular_block(layer_id)
    preset = block.get("preset")
    if preset:
        return str(preset).strip().lower()
    key = (layer_id or "").strip().lower()
    if key == "locspunto":
        return "locspunto"
    if key == "clues":
        return "clues"
    if key in _DENUE_TABULAR_LAYER_IDS:
        return "denue"
    return None


def _layer_has_tabular_capability(layer_id: str) -> bool:
    entry = _raw_layer_entry(layer_id)
    if not entry:
        return False
    if (entry.get("capabilities") or {}).get("tabular"):
        return True
    return _tabular_preset(layer_id) is not None


def list_tabular_layers() -> List[Dict[str, Any]]:
    """Capas habilitadas para el selector de consulta tabular (desde catálogo)."""
    out: List[Dict[str, Any]] = []
    for layer_id in ordered_layer_ids_from_raw():
        if not _layer_has_tabular_capability(layer_id):
            continue
        entry = _raw_layer_entry(layer_id) or {}
        data = entry.get("data") or {}
        table = data.get("table") or T_DENUE
        out.append(
            {
                "id": layer_id,
                "label": entry.get("label") or layer_id,
                "table": table,
            }
        )
    return out


def _resolve_field_specs(
    conn,
    table: str,
    specs: Sequence[Tuple[str, Sequence[str], str]],
) -> List[Dict[str, str]]:
    resolved: List[Dict[str, str]] = []
    for key, candidates, label in specs:
        col = resolve_column(conn, SCHEMA, table, candidates)
        if not col:
            logger.warning("Columna no encontrada en %s: %s", table, candidates)
            continue
        resolved.append({"field": key, "sql": col, "label": label})
    if not resolved:
        raise ValueError("NO_COLUMNS")
    return resolved


def _resolve_clues_columns(conn) -> List[Dict[str, str]]:
    return _resolve_field_specs(conn, T_CLUES, _CLUES_FIELD_SPECS)


def _sql_domicilio_denue(conn, alias: str = "pt") -> str:
    parts: List[str] = []
    for candidates in _DENUE_DOMICILIO_PARTS:
        col = resolve_column(conn, SCHEMA, T_DENUE, candidates)
        if not col:
            continue
        q = f"{alias}.{quote_ident(col)}"
        parts.append(f"NULLIF(TRIM(COALESCE({q}::text, '')), '')")
    if not parts:
        return "NULL::text"
    return f"NULLIF(TRIM(CONCAT_WS(' ', {', '.join(parts)})), '')"


def _resolve_denue_columns(conn) -> List[Dict[str, str]]:
    resolved = _resolve_field_specs(conn, T_DENUE, _DENUE_FIELD_SPECS)
    domicilio_sql = _sql_domicilio_denue(conn, "pt")
    out: List[Dict[str, str]] = []
    for col in resolved:
        out.append(col)
        if col["field"] == "nombre_act":
            out.append({"field": "domicilio", "sql": domicilio_sql, "label": "Domicilio", "computed": True})
    if not any(c["field"] == "domicilio" for c in out):
        out.insert(2, {"field": "domicilio", "sql": domicilio_sql, "label": "Domicilio", "computed": True})
    return out


def _columns_public(columns: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return [{"field": c["field"], "label": c["label"]} for c in columns]


def _columns_with_numero(columns: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return [{"field": NUMERO_FIELD, "label": NUMERO_LABEL}, *_columns_public(columns)]


def _rows_with_numero(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    return [{NUMERO_FIELD: i, **row} for i, row in enumerate(rows, start=1)]


def _build_select_parts(columns: List[Dict[str, str]], alias: str) -> List[str]:
    parts: List[str] = []
    for col in columns:
        expr = col["sql"]
        if col.get("computed"):
            parts.append(f"({expr}) AS {quote_ident(col['field'])}")
        else:
            parts.append(f"{alias}.{quote_ident(col['sql'])} AS {quote_ident(col['field'])}")
    return parts


def _sql_codigo_act_filter(alias: str, codes: Sequence[int]) -> str:
    safe = [int(c) for c in codes if str(c).isdigit()]
    if not safe:
        return "FALSE"
    col = f"regexp_replace(TRIM({alias}.codigo_act::text), '[^0-9]', '', 'g')"
    tests = " OR ".join(f"{col} = '{c}'" for c in safe)
    return f"({tests})"


def _execute_detail_query(
    conn,
    *,
    columns: List[Dict[str, str]],
    select_parts: List[str],
    from_sql: str,
    where_sql: str,
    params: Mapping[str, Any],
    order_by: str,
    with_clause: Optional[str] = None,
) -> Tuple[List[Dict[str, Any]], bool]:
    prefix = f"WITH {with_clause}\n" if with_clause else ""
    sql = f"""
        {prefix}
        SELECT {", ".join(select_parts)}
          FROM {from_sql}
         WHERE {where_sql}
         ORDER BY {order_by}
         LIMIT {MAX_TABULAR_ROWS + 1}
    """
    with conn.cursor() as cur:
        cur.execute(sql, dict(params))
        raw_rows = cur.fetchall()

    truncated = len(raw_rows) > MAX_TABULAR_ROWS
    if truncated:
        raw_rows = raw_rows[:MAX_TABULAR_ROWS]

    rows: List[Dict[str, Any]] = []
    for row in raw_rows:
        item: Dict[str, Any] = {}
        for col in columns:
            key = col["field"]
            item[key] = _serialize_cell(row.get(key) if isinstance(row, dict) else None)
        rows.append(item)
    return rows, truncated


def list_clues_detail_rows(
    conn,
    *,
    where_sql: str,
    params: Mapping[str, Any],
    from_sql: Optional[str] = None,
    with_clause: Optional[str] = None,
) -> Dict[str, Any]:
    """Filas detalle CLUES (con columna No.)."""
    columns = _resolve_clues_columns(conn)
    alias = "pt"
    select_parts = _build_select_parts(columns, alias)
    order_nom = next((c["sql"] for c in columns if c["field"] == "nom_comer"), None)
    order_gid = resolve_column(conn, SCHEMA, T_CLUES, ("gid",)) or columns[0]["sql"]
    order_by = (
        f"LOWER(TRIM(COALESCE({alias}.{quote_ident(order_nom)}::text, ''))) ASC NULLS LAST, "
        f"{alias}.{quote_ident(order_gid)} ASC"
        if order_nom
        else f"{alias}.{quote_ident(order_gid)} ASC"
    )
    rows, truncated = _execute_detail_query(
        conn,
        columns=columns,
        select_parts=select_parts,
        from_sql=from_sql or f"{qualified(T_CLUES)} {alias}",
        where_sql=where_sql,
        params=params,
        order_by=order_by,
        with_clause=with_clause,
    )
    return {
        "columns": _columns_with_numero(columns),
        "rows": _rows_with_numero(rows),
        "filas_truncadas": truncated,
    }


def list_denue_detail_rows(
    conn,
    *,
    codigo_act: Sequence[int],
    where_sql: str,
    params: Mapping[str, Any],
    from_sql: Optional[str] = None,
    with_clause: Optional[str] = None,
    apply_codigo_filter: bool = True,
) -> Dict[str, Any]:
    """Filas detalle DENUE (con columna No. y domicilio concatenado)."""
    columns = _resolve_denue_columns(conn)
    alias = "pt"
    select_parts = _build_select_parts(columns, alias)
    order_nom = next((c["sql"] for c in columns if c["field"] == "nom_estab"), None)
    order_gid = resolve_column(conn, SCHEMA, T_DENUE, ("gid",)) or columns[0]["sql"]
    order_by = (
        f"LOWER(TRIM(COALESCE({alias}.{quote_ident(order_nom)}::text, ''))) ASC NULLS LAST, "
        f"{alias}.{quote_ident(order_gid)} ASC"
        if order_nom
        else f"{alias}.{quote_ident(order_gid)} ASC"
    )
    code_filter = _sql_codigo_act_filter(alias, codigo_act)
    if apply_codigo_filter:
        full_where = f"({where_sql}) AND {code_filter}" if where_sql else code_filter
    else:
        full_where = where_sql
    rows, truncated = _execute_detail_query(
        conn,
        columns=columns,
        select_parts=select_parts,
        from_sql=from_sql or f"{qualified(T_DENUE)} {alias}",
        where_sql=full_where,
        params=params,
        order_by=order_by,
        with_clause=with_clause,
    )
    return {
        "columns": _columns_with_numero(columns),
        "rows": _rows_with_numero(rows),
        "filas_truncadas": truncated,
    }


def tabular_error_message(code: str) -> str:
    return TABULAR_ERRORS.get(code, code)


def _column_specs_from_catalog(layer_id: str) -> Optional[Sequence[Tuple[str, Sequence[str], str]]]:
    block = _tabular_block(layer_id)
    columns = block.get("columns")
    if not isinstance(columns, list) or not columns:
        return None
    specs: List[Tuple[str, Sequence[str], str]] = []
    for col in columns:
        if not isinstance(col, dict):
            continue
        field = str(col.get("field") or "").strip()
        if not field:
            continue
        candidates = col.get("candidates") or [field]
        label = str(col.get("label") or field)
        specs.append((field, tuple(str(c) for c in candidates), label))
    return specs or None


def _resolve_columns_for_layer(conn, layer_id: str, table: str) -> List[Dict[str, str]]:
    specs = _column_specs_from_catalog(layer_id)
    if specs:
        return _resolve_field_specs(conn, table, specs)
    preset = _tabular_preset(layer_id)
    if preset == "locspunto":
        return _resolve_locspunto_columns(conn)
    if preset == "clues":
        return _resolve_clues_columns(conn)
    if preset == "denue":
        return _resolve_denue_columns(conn)
    raise ValueError("NO_COLUMNS")


def fetch_generic_tabular_table(
    conn,
    layer_id: str,
    cve_mun: str,
    *,
    table: str,
    mun_filter_cvegeo: bool = False,
) -> Dict[str, Any]:
    """Consulta tabular genérica con columnas del catálogo o preset legacy."""
    cve = norm_cve_mun(cve_mun)
    if not cve:
        raise ValueError("MISSING_CVE_MUN")

    columns = _resolve_columns_for_layer(conn, layer_id, table)
    alias = "pt"
    select_parts = _build_select_parts(columns, alias)
    where_sql = mun_where_sql(alias, with_cvegeo=mun_filter_cvegeo)
    order_col = columns[0]["sql"]
    order_by = f"{alias}.{quote_ident(order_col)} ASC"

    rows, truncated = _execute_detail_query(
        conn,
        columns=columns,
        select_parts=select_parts,
        from_sql=f"{qualified(table)} {alias}",
        where_sql=where_sql,
        params={"cve": cve},
        order_by=order_by,
    )
    if not rows:
        raise ValueError("NO_ROWS")

    entry = _raw_layer_entry(layer_id) or {}
    nom_mun = _fetch_nom_mun(conn, cve)
    return {
        "layer": layer_id,
        "layer_label": entry.get("label") or layer_id,
        "table": table,
        "cve_mun": cve,
        "nom_mun": nom_mun,
        "total_registros": len(rows),
        "summary_label": f"Registros en el municipio",
        "columns": _columns_with_numero(columns),
        "rows": _rows_with_numero(rows),
        "filas_truncadas": truncated,
    }


def _resolve_locspunto_columns(conn) -> List[Dict[str, str]]:
    resolved: List[Dict[str, str]] = []
    for key, candidates, label in _LOCSPUNTO_FIELD_SPECS:
        col = resolve_column(conn, SCHEMA, T_LOC_PUNTO, candidates)
        if not col:
            logger.warning("Columna no encontrada en %s: %s", T_LOC_PUNTO, candidates)
            continue
        resolved.append({"field": key, "sql": col, "label": label})
    if not resolved:
        raise ValueError("NO_COLUMNS")
    return resolved


def _fetch_nom_mun(conn, cve: str) -> Optional[str]:
    with conn.cursor() as cur:
        cur.execute(
            f"""
            SELECT nomgeo
              FROM {qualified("c_mun")}
             WHERE lpad(right(regexp_replace(TRIM(COALESCE(cve_mun::text, '')), '[^0-9]', '', 'g'), 3), 3, '0') = %(cve)s
             LIMIT 1
            """,
            {"cve": cve},
        )
        row = cur.fetchone()
    if not row:
        return None
    val = row.get("nomgeo") if isinstance(row, dict) else row[0]
    return str(val).strip() if val is not None else None


def _serialize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, (int, float, bool)):
        return value
    return str(value).strip()


def _numeric_sort_expr(sql_col: str) -> str:
    """Orden numérico para claves alfanuméricas (p. ej. cve_loc 0001…9999)."""
    q = quote_ident(sql_col)
    return (
        f"NULLIF(regexp_replace(TRIM(COALESCE({q}::text, '')), '[^0-9]', '', 'g'), '')::bigint"
    )


def fetch_locspunto_table(conn, cve_mun: str) -> Dict[str, Any]:
    cve = norm_cve_mun(cve_mun)
    if not cve:
        raise ValueError("MISSING_CVE_MUN")

    columns = _resolve_locspunto_columns(conn)
    select_parts = [f"{quote_ident(c['sql'])} AS {quote_ident(c['field'])}" for c in columns]
    order_cve_loc = next((c["sql"] for c in columns if c["field"] == "cve_loc"), None)
    order_cvegeo = next((c["sql"] for c in columns if c["field"] == "cvegeo"), columns[0]["sql"])
    if order_cve_loc:
        order_by = (
            f"{_numeric_sort_expr(order_cve_loc)} ASC NULLS LAST, "
            f"{quote_ident(order_cvegeo)} ASC"
        )
    else:
        order_by = f"{quote_ident(order_cvegeo)} ASC"

    sql = f"""
        SELECT {", ".join(select_parts)}
          FROM {qualified(T_LOC_PUNTO)}
         WHERE {mun_where_sql("", with_cvegeo=True)}
         ORDER BY {order_by}
         LIMIT {MAX_TABULAR_ROWS + 1}
    """

    with conn.cursor() as cur:
        cur.execute(sql, {"cve": cve})
        raw_rows = cur.fetchall()

    if len(raw_rows) > MAX_TABULAR_ROWS:
        raw_rows = raw_rows[:MAX_TABULAR_ROWS]

    rows: List[Dict[str, Any]] = []
    for row in raw_rows:
        item: Dict[str, Any] = {}
        for col in columns:
            key = col["field"]
            item[key] = _serialize_cell(row.get(key) if isinstance(row, dict) else None)
        rows.append(item)

    nom_mun = _fetch_nom_mun(conn, cve)
    if not rows:
        raise ValueError("NO_ROWS")

    return {
        "layer": "locspunto",
        "layer_label": (layer_config("locspunto") or {}).get("label", "Localidades"),
        "table": T_LOC_PUNTO,
        "cve_mun": cve,
        "nom_mun": nom_mun,
        "total_registros": len(rows),
        "summary_label": "Localidades en el municipio",
        "columns": [{"field": c["field"], "label": c["label"]} for c in columns],
        "rows": rows,
    }


def fetch_clues_table(conn, cve_mun: str) -> Dict[str, Any]:
    cve = norm_cve_mun(cve_mun)
    if not cve:
        raise ValueError("MISSING_CVE_MUN")

    detail = list_clues_detail_rows(
        conn,
        where_sql=mun_where_sql("pt", with_cvegeo=False),
        params={"cve": cve},
    )
    rows = detail["rows"]
    if not rows:
        raise ValueError("NO_ROWS")

    nom_mun = _fetch_nom_mun(conn, cve)
    return {
        "layer": "clues",
        "layer_label": (layer_config("clues") or {}).get("label", "Establecimientos de salud"),
        "table": T_CLUES,
        "cve_mun": cve,
        "nom_mun": nom_mun,
        "total_registros": len(rows),
        "summary_label": "Establecimientos en el municipio",
        "columns": detail["columns"],
        "rows": rows,
        "filas_truncadas": detail.get("filas_truncadas", False),
    }


def fetch_denue_table(conn, layer_id: str, cve_mun: str) -> Dict[str, Any]:
    cve = norm_cve_mun(cve_mun)
    if not cve:
        raise ValueError("MISSING_CVE_MUN")

    key = (layer_id or "").strip().lower()
    codes = denue_codigos_for_layer(key)
    if not codes:
        raise ValueError("UNKNOWN_LAYER")

    detail = list_denue_detail_rows(
        conn,
        codigo_act=codes,
        where_sql=mun_where_sql("pt", with_cvegeo=False),
        params={"cve": cve},
    )
    rows = detail["rows"]
    if not rows:
        raise ValueError("NO_ROWS")

    nom_mun = _fetch_nom_mun(conn, cve)
    cfg = layer_config(key) or {}
    return {
        "layer": key,
        "layer_label": cfg.get("label") or key,
        "table": T_DENUE,
        "cve_mun": cve,
        "nom_mun": nom_mun,
        "total_registros": len(rows),
        "summary_label": "Establecimientos en el municipio",
        "columns": detail["columns"],
        "rows": rows,
        "filas_truncadas": detail.get("filas_truncadas", False),
    }


def fetch_tabular_data(conn, layer_id: str, cve_mun: str) -> Dict[str, Any]:
    key = (layer_id or "").strip().lower()
    if not _layer_has_tabular_capability(key):
        raise ValueError("UNKNOWN_LAYER")

    preset = _tabular_preset(key)
    block = _tabular_block(key)
    if block.get("columns") or (preset and preset not in _TABULAR_PRESETS):
        entry = _raw_layer_entry(key) or {}
        data = entry.get("data") or {}
        table = str(data.get("table") or T_DENUE)
        mun_cvegeo = data.get("mun_filter_cvegeo") is not False and key == "locspunto"
        return fetch_generic_tabular_table(
            conn,
            key,
            cve_mun,
            table=table,
            mun_filter_cvegeo=mun_cvegeo,
        )

    if preset == "locspunto" or key == "locspunto":
        return fetch_locspunto_table(conn, cve_mun)
    if preset == "clues" or key == "clues":
        return fetch_clues_table(conn, cve_mun)
    if preset == "denue" or key in _DENUE_TABULAR_LAYER_IDS:
        return fetch_denue_table(conn, key, cve_mun)
    raise ValueError("UNKNOWN_LAYER")


def build_tabular_xlsx(payload: Dict[str, Any]) -> bytes:
    """Genera libro Excel con openpyxl (encabezados legibles y resumen municipal)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("EXPORT_FAILED") from exc

    wb = Workbook()
    ws = wb.active
    layer_label = payload.get("layer_label") or "Consulta tabular"
    ws.title = str(layer_label)[:31]

    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    meta_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)
    title_fill = PatternFill("solid", fgColor="0D8A8A")
    header_fill = PatternFill("solid", fgColor="1565C0")
    meta_fill = PatternFill("solid", fgColor="E8F4F8")
    thin = Side(style="thin", color="B0BEC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    columns = payload.get("columns") or []
    rows = payload.get("rows") or []
    nom_mun = payload.get("nom_mun") or ""
    cve_mun = payload.get("cve_mun") or ""
    total = payload.get("total_registros", len(rows))
    summary_label = payload.get("summary_label") or "Total de registros"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(columns), 2))
    title_cell = ws.cell(row=1, column=1, value=f"{layer_label} — {nom_mun} ({cve_mun})")
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    meta_rows = [
        ("Municipio", nom_mun or "—"),
        ("Clave municipal", cve_mun),
        (summary_label, total),
        ("Capa", layer_label),
        ("Tabla origen", f"atlas.{payload.get('table', T_LOC_PUNTO)}"),
    ]
    r = 3
    for label, value in meta_rows:
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        for c in (c1, c2):
            c.font = meta_font if c.column == 1 else body_font
            c.fill = meta_fill
            c.border = border
        r += 1

    header_row = r + 1
    for ci, col in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col.get("label") or col.get("field"))
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 36

    data_row = header_row + 1
    for ri, row in enumerate(rows, start=data_row):
        for ci, col in enumerate(columns, start=1):
            val = row.get(col["field"])
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = body_font
            cell.border = border
            if isinstance(val, (int, float)) and col["field"] != "cvegeo":
                cell.alignment = Alignment(horizontal="right")

    for ci, col in enumerate(columns, start=1):
        letter = get_column_letter(ci)
        max_len = len(str(col.get("label") or ""))
        for row in rows[:200]:
            v = row.get(col["field"])
            if v is not None:
                max_len = max(max_len, len(str(v)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 48)

    ws.freeze_panes = ws.cell(row=data_row, column=1)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
