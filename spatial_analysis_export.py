"""Exportación XLSX del análisis espacial (openpyxl).

Sustituye SheetJS/xlsx en el cliente. El payload es el mismo JSON que
devuelve ``POST /api/analisis/dinamico`` (resultado ya mostrado en el modal).
"""

from __future__ import annotations

import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d")


def _slug(text: str) -> str:
    s = (text or "").strip().lower()
    for a, b in (
        ("á", "a"),
        ("é", "e"),
        ("í", "i"),
        ("ó", "o"),
        ("ú", "u"),
        ("ñ", "n"),
    ):
        s = s.replace(a, b)
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_") or "analisis"


def _fmt_area(value: Any) -> Any:
    if value is None:
        return ""
    try:
        return round(float(value), 2)
    except (TypeError, ValueError):
        return value


def _campo_valor(campo: Dict[str, Any]) -> Any:
    if not isinstance(campo, dict):
        return ""
    val = campo.get("valor")
    return "" if val is None else val


def _append_detail(
    rows: List[List[Any]],
    title: str,
    columns: Optional[Sequence[Dict[str, Any]]],
    detail_rows: Optional[Sequence[Dict[str, Any]]],
) -> None:
    if not columns or not detail_rows:
        return
    rows.append([title])
    rows.append([str(c.get("label") or c.get("field") or "") for c in columns])
    for row in detail_rows:
        rows.append([row.get(c.get("field")) if isinstance(row, dict) else "" for c in columns])
    rows.append([])


def _build_rows(data: Dict[str, Any]) -> List[List[Any]]:
    pol = data.get("poligono") or {}
    modo = data.get("modo")

    if modo == "conteo_multi" and isinstance(data.get("filas"), list):
        sheet: List[List[Any]] = [
            ["Análisis espacial — GroSIG"],
            [],
            ["Área del polígono (m²)", _fmt_area(pol.get("area_m2"))],
            ["Vértices", pol.get("vertices") if pol.get("vertices") is not None else ""],
            [],
            ["Establecimiento", "Total de elementos"],
        ]
        for f in data["filas"]:
            if not isinstance(f, dict):
                continue
            total = f.get("total")
            sheet.append([f.get("etiqueta") or f.get("id") or "", 0 if total is None else total])
        sheet.append([])
        for f in data["filas"]:
            if not isinstance(f, dict):
                continue
            _append_detail(
                sheet,
                f"Detalle — {f.get('etiqueta') or ''}",
                f.get("columns"),
                f.get("rows"),
            )
        return sheet

    if modo == "conteo":
        n = data.get("registros_intersectados")
        if n is None:
            n = 0
        sheet = [
            ["Análisis espacial — GroSIG"],
            [],
            ["Área del polígono (m²)", _fmt_area(pol.get("area_m2"))],
            ["Vértices", pol.get("vertices") if pol.get("vertices") is not None else ""],
            [],
            ["Capa", "Total de elementos"],
            [data.get("capa_etiqueta") or data.get("tabla") or "", n],
            [],
        ]
        _append_detail(
            sheet,
            f"Detalle — {data.get('capa_etiqueta') or ''}",
            data.get("columns"),
            data.get("rows"),
        )
        return sheet

    # agregación
    campos = list(data.get("campos") or [])
    campos.sort(
        key=lambda c: (
            str((c or {}).get("etiqueta") or (c or {}).get("columna") or "").lower()
        )
    )
    sheet = [
        ["Análisis espacial — GroSIG"],
        [],
        ["Área del polígono (m²)", _fmt_area(pol.get("area_m2"))],
        ["Vértices", pol.get("vertices") if pol.get("vertices") is not None else ""],
        ["Capa", data.get("capa_etiqueta") or data.get("tabla") or ""],
        [],
        ["Indicador", "Valor"],
    ]
    for c in campos:
        if not isinstance(c, dict):
            continue
        sheet.append([c.get("etiqueta") or c.get("columna") or "", _campo_valor(c)])
    return sheet


def build_spatial_analysis_xlsx(data: Dict[str, Any]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ValueError("EXPORT_FAILED") from exc

    if not isinstance(data, dict) or not data:
        raise ValueError("RESULTADO_INVALIDO")

    rows = _build_rows(data)
    wb = Workbook()
    ws = wb.active
    ws.title = "Resultados"

    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Calibri", size=11)
    title_fill = PatternFill("solid", fgColor="0D8A8A")
    header_fill = PatternFill("solid", fgColor="1565C0")
    thin = Side(style="thin", color="B0BEC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = body_font
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if r_idx == 1:
                cell.font = title_font
                cell.fill = title_fill
            elif row and str(row[0]) in (
                "Establecimiento",
                "Capa",
                "Indicador",
            ):
                cell.font = header_font
                cell.fill = header_fill

    # Anchos
    max_cols = max((len(r) for r in rows), default=2)
    widths = [56, 18, 28, 20, 16, 24, 24]
    for i in range(1, max_cols + 1):
        ws.column_dimensions[get_column_letter(i)].width = widths[i - 1] if i <= len(widths) else 18

    if max_cols >= 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(max_cols, 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def spatial_export_filename(data: Dict[str, Any]) -> str:
    modo = data.get("modo")
    if modo == "conteo_multi":
        base = "analisis_espacial_denue"
    else:
        base = f"analisis_espacial_{_slug(str(data.get('capa_id') or data.get('tabla') or 'capa'))}"
    return f"{base}_{_stamp()}.xlsx"
