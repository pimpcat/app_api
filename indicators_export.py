"""Exportación CSV / XLSX de indicadores del dashboard (Fase 4).

Genera archivos a partir del payload unificado (``build_indicator_payload``)
y la definición del catálogo (columnas, título, footer).
"""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

from indicators_catalog_loader import indicator_by_id
from indicators_service import IndicatorError, build_indicator_payload

CSV_SEP = ";"


def _slugify(text: str) -> str:
    s = (text or "").strip().lower()
    s = (
        s.replace("á", "a")
        .replace("é", "e")
        .replace("í", "i")
        .replace("ó", "o")
        .replace("ú", "u")
        .replace("ñ", "n")
    )
    s = re.sub(r"[^a-z0-9]+", "_", s)
    return s.strip("_") or "guerrero"


def _stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _selected_label(cve_mun: Optional[str], nom_mun: Optional[str]) -> str:
    cve = (cve_mun or "").strip()
    nom = (nom_mun or "").strip()
    if nom and cve:
        return f"{nom} ({cve})"
    if nom:
        return nom
    if cve:
        return cve
    return "—"


def _metric_keys(ind: Dict[str, Any]) -> List[str]:
    """Columnas de datos a exportar (sin sección/clave/nombre)."""
    exp = ind.get("export") or {}
    skip = {"seccion", "cve_mun", "nom_mun"}
    csv_cols = [
        str(c) for c in (exp.get("csv_columns") or []) if c and str(c) not in skip
    ]
    if csv_cols:
        return csv_cols
    field_keys = [f.get("key") for f in (ind.get("fields") or []) if f.get("key")]
    if field_keys:
        return field_keys
    pres = ind.get("presentation") or {}
    return list(pres.get("bar_metrics") or [])


def _metric_labels(ind: Dict[str, Any], metrics: Sequence[str]) -> List[str]:
    pres = ind.get("presentation") or {}
    legend = pres.get("legend_labels") or []
    by_key = {f.get("key"): f.get("label") or f.get("key") for f in (ind.get("fields") or [])}
    labels: List[str] = []
    for i, key in enumerate(metrics):
        if i < len(legend) and legend[i]:
            labels.append(str(legend[i]))
        else:
            labels.append(str(by_key.get(key) or key))
    return labels


def _title(ind: Dict[str, Any]) -> str:
    pres = ind.get("presentation") or {}
    return str(pres.get("title") or ind.get("label") or ind.get("id") or "Indicador")


def _footer_lines(ind: Dict[str, Any]) -> List[str]:
    pres = ind.get("presentation") or {}
    exp = ind.get("export") or {}
    lines: List[str] = []
    for note in pres.get("notes") or []:
        if note:
            lines.append(str(note))
    footer = pres.get("footer") or exp.get("footer")
    if footer:
        lines.append(str(footer))
    return lines


def _as_row(obj: Dict[str, Any]) -> Dict[str, Any]:
    """Normaliza objetos de entidad/estado a forma cve_mun/nom_mun."""
    row = dict(obj)
    if "cve_mun" not in row or row.get("cve_mun") in (None, ""):
        if obj.get("ent") is not None:
            row["cve_mun"] = obj.get("ent")
    if "nom_mun" not in row or row.get("nom_mun") in (None, ""):
        if obj.get("nom_ent"):
            row["nom_mun"] = obj.get("nom_ent")
        elif obj.get("nomgeo"):
            row["nom_mun"] = obj.get("nomgeo")
    return row


def _section_rows(payload: Dict[str, Any]) -> List[Tuple[str, Dict[str, Any]]]:
    """Filas planas (sección, row) para perfiles ranking / entidad / estados / NSM."""
    out: List[Tuple[str, Dict[str, Any]]] = []

    # Bloques de contexto (nacional / entidad / estatal)
    for key, label in (
        ("tabla_nacional", "Nacional"),
        ("nacional", "Nacional"),
        ("entidad", "Entidad"),
        ("tabla_entidad", "Entidad"),
        ("estatal", "Estatal"),
    ):
        obj = payload.get(key)
        if isinstance(obj, dict) and obj:
            row = _as_row(obj)
            if not row.get("nom_mun"):
                row["nom_mun"] = label
            out.append((label, row))

    n = int(payload.get("ranking_size") or 5)
    top_label = f"Top {n}"
    bottom_label = f"Bottom {n}"

    for r in payload.get("top5") or []:
        tag = f"{top_label} (seleccionado)" if r.get("highlight") else top_label
        out.append((tag, r))
    middle = payload.get("middle")
    if middle:
        row = dict(middle)
        row["highlight"] = True
        out.append(("Seleccionado", row))
    for r in payload.get("bottom5") or []:
        tag = f"{bottom_label} (seleccionado)" if r.get("highlight") else bottom_label
        out.append((tag, r))

    mun = payload.get("municipio")
    if isinstance(mun, dict) and mun:
        out.append(("Municipio", _as_row(mun)))

    for s in payload.get("states") or []:
        if not isinstance(s, dict):
            continue
        out.append(("Entidades", _as_row(s)))

    return out


def _header_row(metrics: Sequence[str], labels: Sequence[str]) -> List[str]:
    return ["Sección", "Clave", "Municipio", *labels]


def _data_row(section: str, row: Dict[str, Any], metrics: Sequence[str]) -> List[Any]:
    cells: List[Any] = [
        section,
        row.get("cve_mun") or "",
        row.get("nom_mun") or "",
    ]
    for key in metrics:
        val = row.get(key)
        cells.append("" if val is None else val)
    return cells


def build_export_table(
    indicator_id: str,
    cve_mun: Optional[str] = None,
    nom_mun: Optional[str] = None,
) -> Dict[str, Any]:
    """Arma tabla tabular + metadatos listos para CSV/XLSX."""
    ind = indicator_by_id(indicator_id)
    if not ind:
        raise IndicatorError("UNKNOWN_INDICATOR", f"Indicador desconocido: {indicator_id}", status=404)

    payload = build_indicator_payload(indicator_id, cve_mun, nom_mun, allow_disabled=True)
    metrics = _metric_keys(ind)
    if not metrics:
        raise IndicatorError("EXPORT_FAILED", "Indicador sin métricas exportables.")

    labels = _metric_labels(ind, metrics)
    header = _header_row(metrics, labels)
    section_rows = _section_rows(payload)
    if not section_rows:
        raise IndicatorError("NO_DATA", "No hay filas para exportar en este indicador.")
    rows = [_data_row(sec, row, metrics) for sec, row in section_rows]

    prefix = (ind.get("export") or {}).get("filename_prefix") or indicator_id
    mun_slug = _slugify(nom_mun or cve_mun or "guerrero")
    filename_base = f"{prefix}_{mun_slug}_{_stamp()}"

    return {
        "indicator_id": indicator_id,
        "title": _title(ind),
        "selected_label": _selected_label(cve_mun, nom_mun),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "footer_lines": _footer_lines(ind),
        "header": header,
        "rows": rows,
        "filename_base": filename_base,
        "metrics": metrics,
        "metric_labels": labels,
    }


def build_indicator_csv(table: Dict[str, Any]) -> bytes:
    """CSV con BOM UTF-8 y separador ``;`` (compatible Excel ES)."""
    buf = io.StringIO(newline="")
    writer = csv.writer(buf, delimiter=CSV_SEP, lineterminator="\r\n", quoting=csv.QUOTE_MINIMAL)
    writer.writerow([f"sep={CSV_SEP}"])
    writer.writerow([table["title"]])
    writer.writerow(["Municipio seleccionado", table["selected_label"]])
    writer.writerow(["Generado", table["generated_at"]])
    writer.writerow([])
    writer.writerow(table["header"])
    for row in table["rows"]:
        writer.writerow(row)
    writer.writerow([])
    for line in table["footer_lines"]:
        writer.writerow([line])
    # BOM + contenido
    return ("\ufeff" + buf.getvalue()).encode("utf-8")


def build_indicator_xlsx(table: Dict[str, Any]) -> bytes:
    """Libro Excel con openpyxl (estilo alineado al export tabular del visor)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise IndicatorError("EXPORT_FAILED", "openpyxl no está disponible en el servidor.") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = str(table.get("title") or "Indicador")[:31]

    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    meta_font = Font(name="Calibri", size=11, bold=True)
    body_font = Font(name="Calibri", size=11)
    title_fill = PatternFill("solid", fgColor="0D8A8A")
    header_fill = PatternFill("solid", fgColor="1565C0")
    meta_fill = PatternFill("solid", fgColor="E8F4F8")
    thin = Side(style="thin", color="B0BEC5")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header = table["header"]
    ncols = max(len(header), 2)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    title_cell = ws.cell(row=1, column=1, value=table["title"])
    title_cell.font = title_font
    title_cell.fill = title_fill
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[1].height = 28

    meta = [
        ("Municipio seleccionado", table["selected_label"]),
        ("Generado", table["generated_at"]),
        ("Indicador", table["indicator_id"]),
    ]
    r = 3
    for label, value in meta:
        c1 = ws.cell(row=r, column=1, value=label)
        c2 = ws.cell(row=r, column=2, value=value)
        c1.font = meta_font
        c2.font = body_font
        for c in (c1, c2):
            c.fill = meta_fill
            c.border = border
        r += 1

    header_row = r + 1
    for ci, col in enumerate(header, start=1):
        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    ws.row_dimensions[header_row].height = 32

    data_row = header_row + 1
    for ri, row in enumerate(table["rows"], start=data_row):
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val if val != "" else None)
            cell.font = body_font
            cell.border = border
            if isinstance(val, (int, float)):
                cell.alignment = Alignment(horizontal="right")

    footer_row = data_row + len(table["rows"]) + 1
    for line in table["footer_lines"]:
        ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=ncols)
        cell = ws.cell(row=footer_row, column=1, value=line)
        cell.font = Font(name="Calibri", size=10, italic=True, color="546E7A")
        footer_row += 1

    for ci in range(1, ncols + 1):
        letter = get_column_letter(ci)
        max_len = len(str(header[ci - 1])) if ci - 1 < len(header) else 8
        for row in table["rows"][:100]:
            if ci - 1 < len(row):
                max_len = max(max_len, len(str(row[ci - 1] if row[ci - 1] is not None else "")))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), 48)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def export_indicator(
    indicator_id: str,
    fmt: str,
    cve_mun: Optional[str] = None,
    nom_mun: Optional[str] = None,
) -> Tuple[bytes, str, str]:
    """Devuelve ``(bytes, filename, media_type)``."""
    fmt_norm = (fmt or "xlsx").strip().lower()
    if fmt_norm in ("excel", "xls"):
        fmt_norm = "xlsx"
    if fmt_norm not in ("csv", "xlsx"):
        raise IndicatorError(
            "INVALID_FORMAT",
            "format debe ser csv o xlsx",
            status=400,
        )

    table = build_export_table(indicator_id, cve_mun, nom_mun)
    base = table["filename_base"]

    if fmt_norm == "csv":
        data = build_indicator_csv(table)
        return data, f"{base}.csv", "text/csv; charset=utf-8"

    data = build_indicator_xlsx(table)
    return (
        data,
        f"{base}.xlsx",
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
